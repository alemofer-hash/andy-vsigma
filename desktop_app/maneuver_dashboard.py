from __future__ import annotations

from itertools import combinations
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from andy_threads.neighbor_detection.config import NeighborDetectionConfig, load_neighborhood_config
from andy_threads.neighbor_detection.dataframe_intake import frame_from_dashboard_agg
from andy_threads.neighbor_detection.feeder_graph import PairDefinition, build_pair_definitions
from andy_threads.neighbor_detection.key_parser import is_feeder_name
from andy_threads.neighbor_detection.models import NeighborManeuverCandidate, NeighborhoodEventEpisode, to_plain_dict
from andy_threads.neighbor_detection.neighborhood_fusion import fuse_candidates
from andy_threads.neighbor_detection.pair_state import analyze_pair
from andy_threads.neighbor_detection.synchronization import synchronize_pair


MANEUVER_SHEET_TITLE = "Manobras"
MANEUVER_DATA_SHEET_TITLE = "Manobras Dados"
CURRENT_VARS = {"IA", "IB", "IC"}
MAX_SELECTION_DERIVED_PAIRS = 45

CLASSIFICATION_LABELS = {
    "LOAD_TRANSFER_CANDIDATE": "Transferência provável de carga",
    "COMMON_MODE_EVENT_CANDIDATE": "Evento comum da subestação",
    "LOCAL_FEEDER_EVENT_CANDIDATE": "Evento local no alimentador",
    "DATA_QUALITY_CANDIDATE": "Qualidade de dados insuficiente",
}

HYPOTHESIS_LABELS = {
    "confirmed_topology_required": "confirmar topologia física",
    "scada_switching_log_required": "confirmar log SCADA ou operação",
    "local_feeder_load_change": "mudança local de carga",
    "measurement_or_cadastre_issue": "possível problema de medição ou cadastro",
    "transfer_balance_low": "balanceamento baixo para transferência",
    "insufficient_common_coverage": "cobertura comum insuficiente",
    "gap_or_missing_values": "lacuna ou valor ausente",
    "source_gaps_present": "existem lacunas na série",
    "substation_common_mode_change": "mudança comum na subestação",
    "possible_upstream_or_common_mode_event": "possível evento a montante ou comum",
}


def normalize_selected_current_variables(values: Iterable[object] | None) -> tuple[str, ...]:
    selected = tuple(
        dict.fromkeys(
            str(value or "").strip().upper()
            for value in values or ()
            if str(value or "").strip().upper() in CURRENT_VARS
        )
    )
    return selected


def resolve_neighborhood_config_path(settings: dict[str, object] | None = None) -> Path | None:
    settings = settings or {}
    for key in ("neighbor_detection_config_path", "feeder_neighborhood_config_path", "andy_threads_feeder_neighborhood_config"):
        raw = str(settings.get(key) or "").strip()
        if raw:
            candidate = Path(raw).expanduser()
            if candidate.exists():
                return candidate
    private_path = Path("config/andy_threads_feeder_neighborhood.private.json")
    if private_path.exists():
        return private_path
    example_path = Path("config/andy_threads_feeder_neighborhood.example.json")
    return example_path if example_path.exists() else None


def detect_dashboard_maneuvers(
    df_agg: pd.DataFrame,
    *,
    selected_variables: Iterable[object] | None,
    selected_feeders: Iterable[object] | None = None,
    settings: dict[str, object] | None = None,
    source_alias: str = "dashboard_xlsx_export",
) -> tuple[list[NeighborManeuverCandidate], list[NeighborhoodEventEpisode], list[str]]:
    current_variables = normalize_selected_current_variables(selected_variables)
    if not current_variables:
        return [], [], ["Manobras não avaliadas: selecione IA, IB ou IC no recorte exportado."]

    config_path = resolve_neighborhood_config_path(settings)
    config = load_neighborhood_config(config_path)
    config = NeighborDetectionConfig(
        substations=config.substations,
        pre_windows_hours=config.pre_windows_hours,
        post_windows_hours=config.post_windows_hours,
        min_persistence_samples=config.min_persistence_samples,
        max_pair_time_offset_intervals=config.max_pair_time_offset_intervals,
        robust_z_threshold=config.robust_z_threshold,
        min_balance_for_transfer=config.min_balance_for_transfer,
        min_quality_score=config.min_quality_score,
        min_common_coverage=config.min_common_coverage,
        current_variables=current_variables,
        paper_reproduction_mode=config.paper_reproduction_mode,
        paper_window_samples=config.paper_window_samples,
        paper_overlap=config.paper_overlap,
    )
    frame, _manifest, _profiles = frame_from_dashboard_agg(
        df_agg,
        source_alias=source_alias,
        current_variables=current_variables,
    )
    if frame.empty:
        return [], [], ["Manobras não avaliadas: recorte sem correntes IA/IB/IC válidas."]
    pair_defs, configured_warnings = build_pair_definitions(frame, config, all_pairs_exploratory=False)
    if not pair_defs:
        pair_defs, selection_warnings = _build_selection_derived_pairs(frame, selected_feeders=selected_feeders)
        if not pair_defs:
            warnings = [*configured_warnings, *selection_warnings]
            warnings.append("Manobras não avaliadas: selecione ao menos dois alimentadores no recorte ou configure a topologia.")
            return [], [], warnings
        warnings = selection_warnings
    else:
        warnings = configured_warnings

    candidates: list[NeighborManeuverCandidate] = []
    run_source_hash = str(frame["source_hash"].dropna().iloc[0]) if "source_hash" in frame and not frame.empty else "dashboard"
    source_name = str(frame["source_alias"].dropna().iloc[0]) if "source_alias" in frame and not frame.empty else source_alias
    for pair_def in pair_defs:
        for variable in current_variables:
            synced = synchronize_pair(
                frame,
                pair_def.se,
                pair_def.first,
                pair_def.second,
                config.min_common_coverage,
                variable=variable,
            )
            result = analyze_pair(
                frame,
                pair_def,
                synced,
                config,
                detector_version="andy_dashboard_current_pair_v0.1",
                source_alias=source_name,
                source_hash=run_source_hash,
            )
            candidates.extend(result.candidates)
    candidates = sorted(candidates, key=lambda item: item.transfer_score, reverse=True)
    cadence_values = [candidate.cadence_seconds for candidate in candidates if candidate.cadence_seconds]
    cadence = float(sorted(cadence_values)[len(cadence_values) // 2]) if cadence_values else None
    episodes = fuse_candidates(candidates, cadence_seconds=cadence)
    return candidates, episodes, warnings


def _build_selection_derived_pairs(
    frame: pd.DataFrame,
    *,
    selected_feeders: Iterable[object] | None,
) -> tuple[list[PairDefinition], list[str]]:
    explicit_selected = _normalize_selected_feeders(selected_feeders)
    present_by_se: dict[str, list[str]] = {
        str(se).upper(): sorted(
            {
                str(value).strip().upper()
                for value in part["feeder"].dropna().unique()
                if is_feeder_name(str(value).strip())
            }
        )
        for se, part in frame.groupby("se")
    }

    pairs: list[PairDefinition] = []
    warnings: list[str] = []
    for se, present in present_by_se.items():
        feeders = [item for item in present if item in explicit_selected] if explicit_selected else present
        if len(feeders) < 2:
            continue
        pair_count = (len(feeders) * (len(feeders) - 1)) // 2
        if pair_count > MAX_SELECTION_DERIVED_PAIRS:
            warnings.append(
                f"Manobras não avaliadas para {se}: seleção geraria {pair_count} pares; reduza o recorte de alimentadores."
            )
            continue
        for first, second in combinations(feeders, 2):
            pairs.append(
                PairDefinition(
                    se=se,
                    first=first,
                    second=second,
                    configured_neighbor_pair=False,
                    exploratory_pair=True,
                )
            )

    if pairs:
        warnings.append(
            "Manobras avaliadas por pares derivados da seleção do engenheiro; confirmar topologia e SCADA antes de validar operacionalmente."
        )
    elif explicit_selected:
        warnings.append("Nenhum par de alimentadores selecionados apareceu com corrente válida no recorte exportado.")
    else:
        warnings.append("Nenhum par configurado; sem seleção explícita de dois alimentadores para derivar pares no dashboard.")
    return pairs, warnings


def _normalize_selected_feeders(values: Iterable[object] | None) -> tuple[str, ...]:
    normalized: list[str] = []
    for value in values or ():
        text = str(value or "").strip().upper()
        if not text:
            continue
        if "|" in text:
            parts = [part.strip().upper() for part in text.split("|")]
            text = parts[1] if len(parts) >= 2 else text
        if is_feeder_name(text) and text not in normalized:
            normalized.append(text)
    return tuple(normalized)


def append_maneuver_dashboard_sheets(
    workbook: Workbook,
    *,
    candidates: list[NeighborManeuverCandidate],
    episodes: list[NeighborhoodEventEpisode],
    warnings: list[str],
    generated_at: str,
) -> None:
    existing = {sheet.title for sheet in workbook.worksheets}
    visible_title = _safe_title(MANEUVER_SHEET_TITLE, existing)
    data_title = _safe_title(MANEUVER_DATA_SHEET_TITLE, existing)
    ws = workbook.create_sheet(visible_title)
    ws_data = workbook.create_sheet(data_title)
    ws.sheet_view.showGridLines = False

    navy = "17324D"
    cyan = "DDEFF6"
    amber = "F7E4B2"
    green = "DDEEDB"
    red = "F2D6D6"
    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=18, color=navy)

    ws["A1"] = "MANOBRAS - Candidatos de transferência entre alimentadores"
    ws["A1"].font = title_font
    ws.merge_cells("A1:L1")
    ws["A2"] = (
        "Leitura operacional: eventos candidatos por salto simultâneo em correntes IA/IB/IC selecionadas. "
        "Nada é confirmado sem validação humana/SCADA."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:L2")

    transfer_count = sum(1 for item in candidates if item.classification.value == "LOAD_TRANSFER_CANDIDATE")
    top = candidates[0] if candidates else None
    cards = [
        ("Candidatos", len(candidates)),
        ("Transferências prováveis", transfer_count),
        ("Fases analisadas", ", ".join(sorted({item.variable for item in candidates})) if candidates else "-"),
        ("Maior score", f"{top.transfer_score:.3f}" if top else "-"),
        ("Direção mais forte", _direction_text(top) if top else "-"),
        ("Gerado em", generated_at),
    ]
    for idx, (label, value) in enumerate(cards):
        col = 1 + (idx % 3) * 4
        row = 4 + (idx // 3) * 3
        ws.cell(row, col, label).font = Font(bold=True, color=navy)
        ws.cell(row, col, label).fill = cyan
        ws.cell(row + 1, col, value).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)

    if warnings:
        ws["A10"] = "Avisos"
        ws["A10"].font = Font(bold=True, color=navy)
        ws["B10"] = " | ".join(str(item) for item in warnings[:4])
        ws["B10"].alignment = Alignment(wrap_text=True)
        ws.merge_cells("B10:L10")

    headers = [
        "Rank",
        "Status",
        "SE",
        "Par",
        "Corrente",
        "Pico",
        "Direção",
        "Delta queda (A)",
        "Delta subida (A)",
        "Balanceamento",
        "Score",
        "Centroide",
        "Hipóteses alternativas",
    ]
    table_row = 12
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(table_row, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rank, candidate in enumerate(candidates, start=1):
        row_idx = table_row + rank
        falling = candidate.feeder_from_candidate or "-"
        rising = candidate.feeder_to_candidate or "-"
        fall_delta = candidate.delta_I_by_feeder.get(falling, 0.0) if falling != "-" else 0.0
        rise_delta = candidate.delta_I_by_feeder.get(rising, 0.0) if rising != "-" else 0.0
        values = [
            rank,
            _classification_label(candidate.classification.value),
            candidate.SE,
            f"{candidate.feeder_pair[0]} x {candidate.feeder_pair[1]}",
            candidate.variable,
            candidate.event_peak,
            _direction_text(candidate),
            fall_delta,
            rise_delta,
            candidate.balance_score,
            candidate.transfer_score,
            candidate.centroid_delta_magnitude,
            _hypotheses_text(candidate.alternative_hypotheses),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "0.000"
        fill = green if candidate.classification.value == "LOAD_TRANSFER_CANDIDATE" else amber
        if candidate.classification.value == "DATA_QUALITY_CANDIDATE":
            fill = red
        ws.cell(row_idx, 2).fill = fill

    for col_idx, width in enumerate([7, 28, 9, 16, 10, 22, 22, 15, 15, 11, 11, 12, 44], start=1):
        ws.column_dimensions[_col_letter(col_idx)].width = width
    ws.freeze_panes = "A13"

    data_headers = [
        "candidate_id",
        "classification",
        "SE",
        "feeder_pair",
        "variable",
        "event_peak",
        "feeder_from_candidate",
        "feeder_to_candidate",
        "delta_I_by_feeder",
        "balance_score",
        "transfer_score",
        "centroid_delta_magnitude",
        "upstream_stability_hint",
        "quality_score",
        "review_status",
        "training_use",
    ]
    for col_idx, header in enumerate(data_headers, start=1):
        cell = ws_data.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, candidate in enumerate(candidates, start=2):
        record = to_plain_dict(candidate)
        data_values = [
            candidate.candidate_id,
            candidate.classification.value,
            candidate.SE,
            " x ".join(candidate.feeder_pair),
            candidate.variable,
            candidate.event_peak,
            candidate.feeder_from_candidate,
            candidate.feeder_to_candidate,
            str(candidate.delta_I_by_feeder),
            candidate.balance_score,
            candidate.transfer_score,
            candidate.centroid_delta_magnitude,
            candidate.upstream_stability_hint,
            candidate.quality_score,
            candidate.review_status.value,
            candidate.training_use,
        ]
        for col_idx, value in enumerate(data_values, start=1):
            ws_data.cell(row_idx, col_idx, value)
        ws_data.cell(row_idx, len(data_headers) + 2, str(record))
    ws_data.sheet_state = "hidden"

    if candidates:
        chart = BarChart()
        chart.title = "Todos os candidatos por score"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Rank"
        max_chart_row = table_row + len(candidates)
        data = Reference(ws, min_col=11, min_row=table_row, max_row=max_chart_row)
        cats = Reference(ws, min_col=1, min_row=table_row + 1, max_row=max_chart_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.height = 7
        chart.width = 18
        ws.add_chart(chart, "A45")

    if not candidates:
        ws["A13"] = "Nenhum candidato de manobra foi detectado no recorte atual."
        ws["A13"].font = Font(bold=True, color=navy)


def _safe_title(desired: str, existing: set[str]) -> str:
    base = "".join(ch for ch in str(desired or "Sheet") if ch not in '[]:*?/\\').strip()[:31] or "Sheet"
    candidate = base
    index = 2
    while candidate in existing:
        suffix = f" ({index})"
        candidate = f"{base[: max(1, 31 - len(suffix))]}{suffix}"
        index += 1
    existing.add(candidate)
    return candidate


def _direction_text(candidate: NeighborManeuverCandidate | None) -> str:
    if candidate is None:
        return "-"
    if candidate.feeder_from_candidate and candidate.feeder_to_candidate:
        return f"{candidate.feeder_from_candidate} -> {candidate.feeder_to_candidate}"
    if candidate.opposite_direction:
        return "sentidos opostos sem direção conclusiva"
    return "sem transferência balanceada"


def _classification_label(value: object) -> str:
    text = str(value or "").strip()
    return CLASSIFICATION_LABELS.get(text, text or "-")


def _hypotheses_text(values: Iterable[object]) -> str:
    labels = [HYPOTHESIS_LABELS.get(str(item), str(item)) for item in values or [] if str(item).strip()]
    return ", ".join(labels) if labels else "-"


def _col_letter(index: int) -> str:
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters
