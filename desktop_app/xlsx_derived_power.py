from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Optional

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from desktop_app.electrical_formulae import DEFAULT_CURRENT_POLICY
from desktop_app.feeder_voltage_resolver import (
    DEFAULT_FEEDER_VLL_KV,
    FeederVoltageLookup,
    iter_feeder_key_candidates,
    normalize_feeder_key,
)
from desktop_app.load_profile_service import _canonical_measurement_name


@dataclass(frozen=True)
class HeaderRows:
    label_row: int = 4
    equip_row: int = 5
    var_row: int = 6
    key_row: int = 7


DEFAULT_HEADER_ROWS = HeaderRows()


@dataclass(frozen=True)
class ElectricalSeriesGroup:
    se: str | None
    bay: str | None
    alimentador: str | None
    equipamento: str | None
    terminal: str | None
    ia_col: int | None
    ib_col: int | None
    ic_col: int | None
    p_col: int | None
    q_col: int | None
    feeder_key: str | None
    vll_kv: float | None
    voltage_source: str | None
    voltage_was_fallback: bool = False


@dataclass(frozen=True)
class DerivedColumnAudit:
    group_key: str
    alimentador: str | None
    normalized_feeder_key: str
    se: str | None
    bay: str | None
    equipamento: str | None
    terminal: str | None
    voltage_kv: float
    voltage_source: str
    voltage_was_fallback: bool
    current_policy: str
    current_columns: str
    q_column: str
    derived_columns: str
    formula_base: str


@dataclass(frozen=True)
class SeriesKeyParts:
    key: str
    se: str
    bay: str
    equipamento: str
    terminal: str
    var: str


def parse_view_mx_series_key(raw_key: object) -> SeriesKeyParts:
    key = str(raw_key or "").strip()
    parts = key.split("|", 4)
    if len(parts) >= 5:
        se, bay, equipamento, terminal, var_name = parts[0], parts[1], parts[2], parts[3], parts[4]
    elif "|" in key:
        se, bay, terminal = "", "", ""
        equipamento, var_name = key.split("|", 1)
    else:
        se, bay, terminal = "", "", ""
        equipamento, var_name = "", key
    return SeriesKeyParts(
        key=key,
        se=str(se).strip(),
        bay=str(bay).strip(),
        equipamento=str(equipamento).strip(),
        terminal=str(terminal).strip(),
        var=str(var_name).strip(),
    )


def build_group_key(
    se: object,
    bay: object,
    equipamento: object,
    terminal: object,
) -> str:
    return "|".join(
        [
            str(se or "").strip(),
            str(bay or "").strip(),
            str(equipamento or "").strip(),
            str(terminal or "").strip(),
        ]
    )


def build_group_metadata_lookup(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if frame.empty:
        return {}

    column_lookup = {normalize_feeder_key(column): str(column) for column in frame.columns}

    def _find_column(*names: str) -> Optional[str]:
        for name in names:
            column_name = column_lookup.get(normalize_feeder_key(name))
            if column_name:
                return column_name
        return None

    se_col = _find_column("SE")
    bay_col = _find_column("BAY")
    equip_col = _find_column("EQUIPAMENTO", "equip_id")
    terminal_col = _find_column("TERMINAL")
    ponto_id_col = _find_column("ponto_id")
    equip_id_col = _find_column("equip_id")
    var_col = _find_column("var")
    alimentador_col = _find_column("alimentador", "FEEDER", "ALIMENTADOR")
    al_col = _find_column("AL")
    circuito_col = _find_column("circuito", "circuit")

    lookup: dict[str, dict[str, Any]] = {}
    for row in frame.to_dict(orient="records"):
        se_value = row.get(se_col) if se_col else ""
        bay_value = row.get(bay_col) if bay_col else ""
        equipamento_value = row.get(equip_col) if equip_col else ""
        terminal_value = row.get(terminal_col) if terminal_col else ""
        var_value = row.get(var_col) if var_col else ""

        series_key = build_group_key(se_value, bay_value, equipamento_value, terminal_value)
        if var_value:
            series_key = f"{series_key}|{str(var_value).strip()}"
        group_key = build_group_key(se_value, bay_value, equipamento_value, terminal_value)
        metadata = {
            "SE": se_value,
            "BAY": bay_value,
            "EQUIPAMENTO": equipamento_value,
            "TERMINAL": terminal_value,
            "equip_id": row.get(equip_id_col) if equip_id_col else "",
            "ponto_id": row.get(ponto_id_col) if ponto_id_col else "",
            "alimentador": row.get(alimentador_col) if alimentador_col else "",
            "AL": row.get(al_col) if al_col else "",
            "circuito": row.get(circuito_col) if circuito_col else "",
        }
        if group_key and group_key not in lookup:
            lookup[group_key] = metadata
        if series_key and series_key not in lookup:
            lookup[series_key] = metadata
    return lookup


def _copy_cell_style(source_cell, target_cell) -> None:
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)


def _resolve_voltage_entry(
    *,
    feeder_key: str | None,
    metadata: Mapping[str, Any] | None,
    voltage_entries: Mapping[str, FeederVoltageLookup],
) -> tuple[float, str, str, bool]:
    for normalized_candidate in iter_feeder_key_candidates(feeder_key=feeder_key, metadata=metadata):
        entry = voltage_entries.get(normalized_candidate)
        if entry is not None:
            return entry.voltage_kv, entry.source, normalized_candidate, False
    normalized_fallback = normalize_feeder_key(feeder_key or (metadata or {}).get("BAY") or "")
    return DEFAULT_FEEDER_VLL_KV, "fallback:23.1kV", normalized_fallback, True


def discover_electrical_series_groups(
    ws: Worksheet,
    header_model: HeaderRows = DEFAULT_HEADER_ROWS,
    point_metadata: Mapping[str, Mapping[str, Any]] | None = None,
    voltage_entries: Mapping[str, FeederVoltageLookup] | None = None,
    options: Any | None = None,
) -> list[ElectricalSeriesGroup]:
    _ = options
    voltage_lookup = voltage_entries or {}
    groups: dict[str, dict[str, Any]] = {}

    for column_idx in range(1, ws.max_column + 1):
        raw_key = ws.cell(header_model.key_row, column_idx).value
        if not raw_key or str(raw_key).strip() in {"KEY", "-"}:
            continue

        series_key = parse_view_mx_series_key(raw_key)
        canonical_var = _canonical_measurement_name(ws.cell(header_model.var_row, column_idx).value) or _canonical_measurement_name(series_key.var)
        if canonical_var not in {"IA", "IB", "IC", "P", "Q"}:
            continue

        group_key = build_group_key(series_key.se, series_key.bay, series_key.equipamento, series_key.terminal)
        metadata = (
            (point_metadata or {}).get(series_key.key)
            or (point_metadata or {}).get(group_key)
            or {
                "SE": series_key.se,
                "BAY": series_key.bay,
                "EQUIPAMENTO": series_key.equipamento,
                "TERMINAL": series_key.terminal,
                "ponto_id": series_key.key,
            }
        )
        feeder_key = (
            str(metadata.get("alimentador") or "").strip()
            or str(metadata.get("AL") or "").strip()
            or str(metadata.get("BAY") or metadata.get("bay") or "").strip()
            or series_key.bay
        )
        vll_kv, voltage_source, normalized_feeder_key, voltage_was_fallback = _resolve_voltage_entry(
            feeder_key=feeder_key,
            metadata=metadata,
            voltage_entries=voltage_lookup,
        )

        bucket = groups.setdefault(
            group_key,
            {
                "se": series_key.se,
                "bay": series_key.bay,
                "alimentador": str(metadata.get("alimentador") or metadata.get("AL") or series_key.bay or "").strip() or None,
                "equipamento": series_key.equipamento,
                "terminal": series_key.terminal,
                "ia_col": None,
                "ib_col": None,
                "ic_col": None,
                "p_col": None,
                "q_col": None,
                "feeder_key": feeder_key or None,
                "vll_kv": vll_kv,
                "voltage_source": voltage_source,
                "voltage_was_fallback": voltage_was_fallback,
                "normalized_feeder_key": normalized_feeder_key,
            },
        )
        bucket[f"{canonical_var.lower()}_col"] = column_idx

    discovered: list[ElectricalSeriesGroup] = []
    for bucket in groups.values():
        if not any(bucket.get(name) for name in ("ia_col", "ib_col", "ic_col")):
            continue
        if not bucket.get("q_col"):
            continue
        discovered.append(
            ElectricalSeriesGroup(
                se=bucket.get("se"),
                bay=bucket.get("bay"),
                alimentador=bucket.get("alimentador"),
                equipamento=bucket.get("equipamento"),
                terminal=bucket.get("terminal"),
                ia_col=bucket.get("ia_col"),
                ib_col=bucket.get("ib_col"),
                ic_col=bucket.get("ic_col"),
                p_col=bucket.get("p_col"),
                q_col=bucket.get("q_col"),
                feeder_key=bucket.get("feeder_key"),
                vll_kv=float(bucket.get("vll_kv") or DEFAULT_FEEDER_VLL_KV),
                voltage_source=str(bucket.get("voltage_source") or ""),
                voltage_was_fallback=bool(bucket.get("voltage_was_fallback")),
            )
        )

    discovered.sort(
        key=lambda item: (
            str(item.se or ""),
            str(item.bay or ""),
            str(item.equipamento or ""),
            str(item.terminal or ""),
        )
    )
    return discovered


def _first_source_column(group: ElectricalSeriesGroup) -> Optional[int]:
    return group.ia_col or group.ib_col or group.ic_col or group.q_col or group.p_col


def _build_current_formula_expr(
    row_idx: int,
    group: ElectricalSeriesGroup,
    current_policy: str,
) -> Optional[str]:
    phase_columns = [group.ia_col, group.ib_col, group.ic_col]
    phase_refs = [f"{get_column_letter(column_idx)}{row_idx}" for column_idx in phase_columns if column_idx]
    if not phase_refs:
        return None

    normalized_policy = str(current_policy or DEFAULT_CURRENT_POLICY).strip().lower()
    if normalized_policy == "first_available_phase":
        return f"ABS({phase_refs[0]})"
    if len(phase_refs) == 1:
        return f"ABS({phase_refs[0]})"
    joined_refs = ",".join(f"ABS({cell_ref})" for cell_ref in phase_refs)
    return f"MAX({joined_refs})"


def append_derived_power_columns_to_view_mx(
    ws: Worksheet,
    groups: list[ElectricalSeriesGroup],
    first_data_row: int,
    last_data_row: int,
    header_rows: HeaderRows = DEFAULT_HEADER_ROWS,
    current_policy: str = DEFAULT_CURRENT_POLICY,
) -> list[DerivedColumnAudit]:
    audits: list[DerivedColumnAudit] = []
    next_column = ws.max_column + 1

    for group in groups:
        template_column = _first_source_column(group)
        if template_column is None or group.q_col is None or group.vll_kv is None:
            continue

        group_key = build_group_key(group.se, group.bay, group.equipamento, group.terminal)
        derived_headers = [
            (next_column, "MVA", f"{group_key}|MVA_CALC"),
            (next_column + 1, "FP", f"{group_key}|FP_CALC"),
            (next_column + 2, "MW", f"{group_key}|MW_CALC"),
        ]

        for column_idx, visible_var, key_text in derived_headers:
            label_cell = ws.cell(header_rows.label_row, column_idx, key_text)
            equip_cell = ws.cell(header_rows.equip_row, column_idx, group.equipamento or "")
            var_cell = ws.cell(header_rows.var_row, column_idx, visible_var)
            key_cell = ws.cell(header_rows.key_row, column_idx, key_text)
            for source_row, target_cell in (
                (header_rows.label_row, label_cell),
                (header_rows.equip_row, equip_cell),
                (header_rows.var_row, var_cell),
                (header_rows.key_row, key_cell),
            ):
                _copy_cell_style(ws.cell(source_row, template_column), target_cell)

        mva_col, fp_col, mw_col = next_column, next_column + 1, next_column + 2
        mva_letter = get_column_letter(mva_col)
        fp_letter = get_column_letter(fp_col)
        mw_letter = get_column_letter(mw_col)
        q_letter = get_column_letter(group.q_col)
        voltage_literal = f"{float(group.vll_kv):.6f}".rstrip("0").rstrip(".")

        for row_idx in range(first_data_row, last_data_row + 1):
            source_style_cell = ws.cell(row_idx, template_column)
            current_expr = _build_current_formula_expr(row_idx=row_idx, group=group, current_policy=current_policy)
            if current_expr is None:
                continue

            mva_ref = f"{mva_letter}{row_idx}"
            fp_ref = f"{fp_letter}{row_idx}"
            q_ref = f"{q_letter}{row_idx}"

            mva_cell = ws.cell(row_idx, mva_col)
            fp_cell = ws.cell(row_idx, fp_col)
            mw_cell = ws.cell(row_idx, mw_col)
            for target_cell in (mva_cell, fp_cell, mw_cell):
                _copy_cell_style(source_style_cell, target_cell)

            mva_cell.value = f'=IFERROR({current_expr}*SQRT(3)*{voltage_literal}/1000,"")'
            fp_cell.value = (
                f'=IFERROR(IF(OR({mva_ref}="",{mva_ref}=0),"",SQRT(MAX(0,1-({q_ref}/{mva_ref})^2))),"")'
            )
            mw_cell.value = f'=IFERROR(IF(OR({mva_ref}="",{mva_ref}=0,{fp_ref}=""),"",{mva_ref}*{fp_ref}),"")'
            mva_cell.number_format = "0.000"
            fp_cell.number_format = "0.0000"
            mw_cell.number_format = "0.000"

        ws.column_dimensions[mva_letter].width = 12
        ws.column_dimensions[fp_letter].width = 10
        ws.column_dimensions[mw_letter].width = 12

        current_columns = ",".join(
            get_column_letter(column_idx)
            for column_idx in (group.ia_col, group.ib_col, group.ic_col)
            if column_idx is not None
        )
        audits.append(
            DerivedColumnAudit(
                group_key=group_key,
                alimentador=group.alimentador or group.feeder_key,
                normalized_feeder_key=normalize_feeder_key(group.feeder_key or group.alimentador or group.bay or ""),
                se=group.se,
                bay=group.bay,
                equipamento=group.equipamento,
                terminal=group.terminal,
                voltage_kv=float(group.vll_kv),
                voltage_source=str(group.voltage_source or ""),
                voltage_was_fallback=bool(group.voltage_was_fallback),
                current_policy=current_policy,
                current_columns=current_columns,
                q_column=q_letter,
                derived_columns=f"{mva_letter},{fp_letter},{mw_letter}",
                formula_base=f"S=sqrt(3)*Vll*I/1000; FP=sqrt(1-(Q/S)^2); MW=S*FP; Vll={voltage_literal}kV",
            )
        )
        next_column += 3

    return audits


def append_derived_power_audit_section(
    ws_audit: Worksheet,
    audits: list[DerivedColumnAudit],
    *,
    export_timestamp: str,
    als_reference_path: Path | None,
) -> None:
    if not audits:
        return

    start_row = ws_audit.max_row + 3 if ws_audit.max_row else 1
    ws_audit.cell(start_row, 1, "POWER_DERIVED_COLUMNS_AUDIT")
    ws_audit.cell(start_row + 1, 1, "ALS_REFERENCE_PATH")
    ws_audit.cell(start_row + 1, 2, str(als_reference_path) if als_reference_path else "(fallback_only)")

    headers = [
        "EXPORT_TIMESTAMP",
        "ALIMENTADOR",
        "KEY",
        "SE",
        "BAY",
        "EQUIPAMENTO",
        "TERMINAL",
        "V_LL_KV",
        "SOURCE",
        "FALLBACK",
        "CURRENT_POLICY",
        "I_COLUMNS",
        "Q_COLUMN",
        "DERIVED_COLUMNS",
        "FORMULA_BASE",
    ]
    header_row = start_row + 3
    template_header_cell = ws_audit.cell(1, 1)
    for column_idx, header in enumerate(headers, start=1):
        cell = ws_audit.cell(header_row, column_idx, header)
        _copy_cell_style(template_header_cell, cell)

    for row_offset, audit in enumerate(audits, start=1):
        row_idx = header_row + row_offset
        values = [
            export_timestamp,
            audit.alimentador or "",
            audit.group_key,
            audit.se or "",
            audit.bay or "",
            audit.equipamento or "",
            audit.terminal or "",
            audit.voltage_kv,
            audit.voltage_source,
            bool(audit.voltage_was_fallback),
            audit.current_policy,
            audit.current_columns,
            audit.q_column,
            audit.derived_columns,
            audit.formula_base,
        ]
        for column_idx, value in enumerate(values, start=1):
            ws_audit.cell(row_idx, column_idx, value)

    ws_audit.column_dimensions["A"].width = max(ws_audit.column_dimensions["A"].width or 0, 20)
    ws_audit.column_dimensions["B"].width = max(ws_audit.column_dimensions["B"].width or 0, 20)
    ws_audit.column_dimensions["C"].width = max(ws_audit.column_dimensions["C"].width or 0, 38)
    ws_audit.column_dimensions["H"].width = max(ws_audit.column_dimensions["H"].width or 0, 12)
    ws_audit.column_dimensions["I"].width = max(ws_audit.column_dimensions["I"].width or 0, 28)
