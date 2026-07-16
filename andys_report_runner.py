from __future__ import annotations

import argparse
import datetime as dt
import math
import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import duckdb
import openpyxl
import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from config import DEFAULT_SOURCE_ROOT, get_db_path, get_default_work_root, get_export_dir
from export_formatting import (
    TIMESTAMP_SPLIT_DATE_COLUMN,
    TIMESTAMP_SPLIT_EXCEL_DATE_NUMBER_FORMAT,
    TIMESTAMP_SPLIT_EXCEL_TIME_NUMBER_FORMAT,
    TIMESTAMP_SPLIT_TIME_COLUMN,
    neutralize_formula_value,
)
from measurement_value import normalize_measurement_series
from utils.formatting import round_half_up_float


DEFAULT_WORK_ROOT = get_default_work_root()
DEFAULT_DB = get_db_path(DEFAULT_WORK_ROOT)
DEFAULT_OUT_DIR = get_export_dir(DEFAULT_WORK_ROOT)

ENV_LABEL = "Andy's"
TS_FMT = "%Y-%m-%d %H:%M:%S"
TS_PLOT_FMT = "%d/%m/%Y %H:%M"
EXCEL_MAX_ROWS = 1_048_576
LARGE_ROWS_WARN = 1_500_000
AUTO_TIMEFLOOR_ROWS = 3_000_000
PLOT_DATA_SHEET = "PLOT_DATA"
DASHBOARD_CHART_START_ROW = 12
DASHBOARD_CHART_START_COL = 2  # B
DASHBOARD_CHART_END_COL = 13  # M
DASHBOARD_CHART_ROW_GAP = 3
DASHBOARD_CHART_ROW_HEIGHT_PT = 18
DASHBOARD_AXIS_NUMFMT = "+0.0;-0.0;0.0"
TIME_AUDIT_SHEET = "TIME_AUDIT"


def log(prefix: str, msg: str) -> None:
    print(f"[{prefix}] {msg}")


def _ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


def sql_quote(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


@dataclass(frozen=True)
class DashboardSeriesSpec:
    key: str
    se: str
    bay: str
    equip: str
    terminal: str
    var: str


@dataclass(frozen=True)
class DashboardChartGroup:
    title: str
    series: List[DashboardSeriesSpec]


@dataclass(frozen=True)
class DashboardSeriesAuditEntry:
    key: str
    label: str
    plot_column: int
    plotted: bool
    chart_title: str


@dataclass(frozen=True)
class DashboardAxisSpec:
    min_value: Optional[float]
    max_value: Optional[float]
    crosses_mode: str
    num_format: str
    has_negative: bool
    has_positive: bool


@dataclass(frozen=True)
class DashboardChartLayoutSpec:
    title: str
    series_count: int
    legend_position: str
    anchor_from_row: int
    anchor_to_row: int
    anchor_from_col: int
    anchor_to_col: int
    axis: DashboardAxisSpec


SHIFT_ORDER = ("Madrugada", "Manha", "Tarde", "Noite")


@dataclass(frozen=True)
class PatamarShiftStats:
    shift: str
    p_min: float
    p_max: float
    q_min: float
    q_max: float
    p_text: str
    q_text: str


@dataclass(frozen=True)
class PatamarResult:
    shifts: List[PatamarShiftStats]
    p_key_used: str
    q_key_used: str
    warnings: List[str]


@dataclass(frozen=True)
class PatamarWriteSpec:
    summary_sheet: str = "PATAMAR - Resumo"
    first_row: int = 6
    p_column: str = "L"
    q_column: str = "M"


def parse_dashboard_series_key(raw_key: object) -> DashboardSeriesSpec:
    key = str(raw_key or "").strip()
    parts = key.split("|", 4)
    if len(parts) >= 5:
        se, bay, equip, terminal, var = parts[0], parts[1], parts[2], parts[3], parts[4]
    elif "|" in key:
        se, bay, terminal = "", "", ""
        equip, var = key.split("|", 1)
    else:
        se, bay, terminal = "", "", ""
        equip, var = "", key
    return DashboardSeriesSpec(
        key=key,
        se=str(se).strip(),
        bay=str(bay).strip(),
        equip=str(equip).strip(),
        terminal=str(terminal).strip(),
        var=str(var).strip(),
    )


def build_dashboard_series_label(spec: DashboardSeriesSpec) -> str:
    equip_txt = spec.equip or "-"
    var_txt = spec.var or "-"
    se_txt = spec.se or "-"
    bay_txt = spec.bay or "-"
    terminal_txt = spec.terminal or "-"
    return str(neutralize_formula_value(f"{equip_txt} | {var_txt} | SE={se_txt} | BAY={bay_txt} | T={terminal_txt}"))


def resolve_dashboard_series_specs(
    wide_columns: List[object],
    *,
    selected_pairs: Optional[List[Tuple[str, str]]] = None,
) -> List[DashboardSeriesSpec]:
    specs: List[DashboardSeriesSpec] = []
    wanted_pairs = None
    if selected_pairs:
        wanted_pairs = {
            (str(eq).strip(), str(vv).strip())
            for eq, vv in selected_pairs
            if str(eq).strip() and str(vv).strip()
        }
    seen_keys = set()
    for raw_key in wide_columns:
        spec = parse_dashboard_series_key(raw_key)
        if not spec.key or spec.key in seen_keys:
            continue
        if wanted_pairs is not None and (spec.equip, spec.var) not in wanted_pairs:
            continue
        specs.append(spec)
        seen_keys.add(spec.key)
    specs.sort(key=lambda item: (item.var, item.equip, item.bay, item.terminal, item.se, item.key))
    return specs


def build_dashboard_chart_groups(
    specs: List[DashboardSeriesSpec],
    *,
    max_series_per_chart: int,
) -> List[DashboardChartGroup]:
    if not specs:
        return []
    groups: List[DashboardChartGroup] = []
    per_chart = max(1, int(max_series_per_chart))
    if len(specs) <= per_chart:
        return [DashboardChartGroup(title=f"Series exportadas | {len(specs)} serie(s)", series=list(specs))]
    by_var: Dict[str, List[DashboardSeriesSpec]] = {}
    for spec in specs:
        by_var.setdefault(spec.var or "-", []).append(spec)
    for var_name, var_specs in by_var.items():
        total_parts = max(1, (len(var_specs) + per_chart - 1) // per_chart)
        for idx in range(total_parts):
            chunk = var_specs[idx * per_chart : (idx + 1) * per_chart]
            title = f"VAR={var_name or '-'} | {len(chunk)} serie(s)"
            if total_parts > 1:
                title += f" | parte {idx + 1}/{total_parts}"
            groups.append(DashboardChartGroup(title=title, series=list(chunk)))
    return groups


def build_dashboard_series_audit(
    specs: List[DashboardSeriesSpec],
    chart_groups: List[DashboardChartGroup],
) -> List[DashboardSeriesAuditEntry]:
    key_to_chart: Dict[str, str] = {}
    for group in chart_groups:
        for spec in group.series:
            key_to_chart[spec.key] = group.title
    entries: List[DashboardSeriesAuditEntry] = []
    for idx, spec in enumerate(specs, start=2):
        chart_title = key_to_chart.get(spec.key, "")
        entries.append(
            DashboardSeriesAuditEntry(
                key=spec.key,
                label=build_dashboard_series_label(spec),
                plot_column=idx,
                plotted=bool(chart_title),
                chart_title=chart_title,
            )
        )
    return entries


def resolve_dashboard_legend_position(series_count: int) -> str:
    return "r" if int(series_count) <= 4 else "b"


def resolve_dashboard_chart_rowspan(series_count: int, legend_position: str) -> int:
    count = max(1, int(series_count))
    if legend_position == "r":
        return 16 + max(0, math.ceil(count / 4) - 1) * 2
    return 18 + max(0, math.ceil(count / 4)) * 2


def build_dashboard_axis_spec(values: pd.Series) -> DashboardAxisSpec:
    numeric = pd.to_numeric(values, errors="coerce").dropna()
    if numeric.empty:
        return DashboardAxisSpec(
            min_value=None,
            max_value=None,
            crosses_mode="min",
            num_format=DASHBOARD_AXIS_NUMFMT,
            has_negative=False,
            has_positive=False,
        )

    data_min = float(numeric.min())
    data_max = float(numeric.max())
    span = float(data_max - data_min)
    magnitude = max(abs(data_min), abs(data_max), 1.0)
    if span <= 0:
        pad = max(magnitude * 0.12, 0.5)
    else:
        pad = max(span * 0.08, magnitude * 0.03, 0.2)

    axis_min = data_min - pad
    axis_max = data_max + pad
    has_negative = bool(data_min < 0)
    has_positive = bool(data_max > 0)
    crosses_zero = bool(data_min <= 0 <= data_max)
    if crosses_zero:
        axis_min = min(axis_min, 0.0)
        axis_max = max(axis_max, 0.0)
        crosses_mode = "autoZero"
    elif data_max < 0:
        crosses_mode = "max"
    else:
        crosses_mode = "min"

    return DashboardAxisSpec(
        min_value=round_half_up_float(axis_min, ndigits=2),
        max_value=round_half_up_float(axis_max, ndigits=2),
        crosses_mode=crosses_mode,
        num_format=DASHBOARD_AXIS_NUMFMT,
        has_negative=has_negative,
        has_positive=has_positive,
    )


def build_dashboard_chart_layouts(
    chart_groups: List[DashboardChartGroup],
    plot_frame: pd.DataFrame,
    *,
    start_row: int = DASHBOARD_CHART_START_ROW,
    start_col: int = DASHBOARD_CHART_START_COL,
    end_col: int = DASHBOARD_CHART_END_COL,
) -> List[DashboardChartLayoutSpec]:
    plans: List[DashboardChartLayoutSpec] = []
    next_row = int(start_row)
    for group in chart_groups:
        legend_position = resolve_dashboard_legend_position(len(group.series))
        row_span = resolve_dashboard_chart_rowspan(len(group.series), legend_position)
        group_columns = [spec.key for spec in group.series if spec.key in plot_frame.columns]
        group_values = pd.Series(dtype="float64")
        if group_columns:
            group_values = pd.Series(plot_frame[group_columns].to_numpy().ravel())
        axis = build_dashboard_axis_spec(group_values)
        plan = DashboardChartLayoutSpec(
            title=group.title,
            series_count=len(group.series),
            legend_position=legend_position,
            anchor_from_row=next_row,
            anchor_to_row=next_row + row_span - 1,
            anchor_from_col=int(start_col),
            anchor_to_col=int(end_col),
            axis=axis,
        )
        plans.append(plan)
        next_row = plan.anchor_to_row + DASHBOARD_CHART_ROW_GAP + 1
    return plans


def build_temporal_gap_audit(ts_series: pd.Series) -> Tuple[pd.DataFrame, Dict[str, object]]:
    ts = pd.to_datetime(ts_series, errors="coerce").dropna().sort_values().drop_duplicates()
    if ts.empty:
        return (
            pd.DataFrame(columns=["prev_ts", "curr_ts", "delta_minutes", "expected_minutes", "missing_intervals", "gap_kind"]),
            {"status": "empty", "expected_minutes": 0.0, "gap_count": 0, "first_gap_after": ""},
        )

    frame = pd.DataFrame({"curr_ts": ts})
    frame["prev_ts"] = frame["curr_ts"].shift(1)
    frame["delta_minutes"] = (frame["curr_ts"] - frame["prev_ts"]).dt.total_seconds() / 60.0
    deltas = frame["delta_minutes"].dropna()
    expected_minutes = float(deltas.mode().iloc[0]) if not deltas.empty else 0.0
    if expected_minutes > 0:
        gaps = frame[frame["delta_minutes"] > (expected_minutes + 1e-9)].copy()
        gaps["missing_intervals"] = (
            (gaps["delta_minutes"] / expected_minutes).round().astype(int) - 1
        ).clip(lower=1)
    else:
        gaps = frame.iloc[0:0].copy()
        gaps["missing_intervals"] = pd.Series(dtype="int64")
    gaps["expected_minutes"] = expected_minutes
    gaps["gap_kind"] = gaps["delta_minutes"].map(
        lambda value: "day_gap" if float(value or 0.0) >= 24.0 * 60.0 else "incomplete_window"
    )
    return (
        gaps[["prev_ts", "curr_ts", "delta_minutes", "expected_minutes", "missing_intervals", "gap_kind"]],
        {
            "status": "ok" if gaps.empty else "gap_detected",
            "expected_minutes": expected_minutes,
            "gap_count": int(len(gaps)),
            "first_gap_after": gaps["prev_ts"].iloc[0].strftime(TS_FMT) if not gaps.empty else "",
        },
    )


def chart_anchor_label(plan: DashboardChartLayoutSpec) -> str:
    from_col = openpyxl.utils.get_column_letter(plan.anchor_from_col)
    to_col = openpyxl.utils.get_column_letter(plan.anchor_to_col)
    return f"{from_col}{plan.anchor_from_row}:{to_col}{plan.anchor_to_row}"


def add_named_range(wb, name, sheet, col_letter, n):
    if n <= 0:
        text = f"{sheet}!${col_letter}$2:${col_letter}$2"
    else:
        text = f"{sheet}!${col_letter}$2:${col_letter}${n+1}"
    wb.defined_names.add(DefinedName(name=name, attr_text=text))


def validate_ts_or_raise(ts: str, field_name: str) -> dt.datetime:
    raw = (ts or "").strip()
    try:
        return dt.datetime.strptime(raw, TS_FMT)
    except ValueError as e:
        raise ValueError(
            f"Timestamp invalido em '{field_name}': {raw!r}. "
            f"Use o formato: {TS_FMT}"
        ) from e


def connect_duckdb(db_path: str) -> duckdb.DuckDBPyConnection:
    if not os.path.exists(db_path):
        raise FileNotFoundError(
            f"Nao encontrei o DuckDB em: {db_path}. "
            f"O ANDY usa por padrao a origem corporativa {DEFAULT_SOURCE_ROOT}. "
            "Se a rede/share estiver indisponivel, conecte a VPN ou a rede corporativa e rode a indexacao novamente."
        )
    con = duckdb.connect(db_path, read_only=True)
    con.execute("PRAGMA threads=4;")
    return con


def ensure_medicoes_exists(con: duckdb.DuckDBPyConnection) -> None:
    row = con.execute(
        """
        SELECT table_type
        FROM information_schema.tables
        WHERE table_schema = 'main' AND table_name = 'medicoes'
        LIMIT 1;
        """
    ).fetchone()
    if row is None:
        raise RuntimeError(
            "Tabela/view 'medicoes' nao encontrada no DuckDB. "
            "Rode o andys_indexer.py para criar o lake e o catalogo."
        )
    log("LAKE", f"Objeto 'medicoes' encontrado ({row[0]}).")


def lake_overview(con: duckdb.DuckDBPyConnection) -> Dict[str, str]:
    row = con.execute(
        """
        SELECT
          COUNT(*)::BIGINT AS rows_total,
          MIN(timestamp) AS ts_min,
          MAX(timestamp) AS ts_max,
          COUNT(DISTINCT equip_id) AS equips_distintos,
          COUNT(DISTINCT var) AS vars_distintas
        FROM medicoes;
        """
    ).fetchone()
    return {
        "rows_total": str(row[0]),
        "ts_min": str(row[1]),
        "ts_max": str(row[2]),
        "equips_distintos": str(row[3]),
        "vars_distintas": str(row[4]),
    }


def build_where_sql(
    equips: List[str],
    t0: str,
    t1: str,
    vars_: Optional[List[str]],
) -> str:
    if not equips:
        raise ValueError("Informe pelo menos 1 equipamento via --equip.")
    equips_sql = "(" + ",".join([sql_quote(e) for e in equips]) + ")"
    clauses = [
        f"equip_id IN {equips_sql}",
        f"timestamp >= TIMESTAMP {sql_quote(t0)}",
        f"timestamp <= TIMESTAMP {sql_quote(t1)}",
    ]
    if vars_:
        vars_sql = "(" + ",".join([sql_quote(v) for v in vars_]) + ")"
        clauses.append(f"var IN {vars_sql}")
    return " AND ".join(clauses)


def count_recorte(
    con: duckdb.DuckDBPyConnection,
    where_sql: str,
) -> int:
    return int(con.execute(f"SELECT COUNT(*) FROM medicoes WHERE {where_sql};").fetchone()[0])


def fetch_recorte_long(
    con: duckdb.DuckDBPyConnection,
    where_sql: str,
) -> pd.DataFrame:
    q = f"""
        SELECT
          timestamp,
          equip_id,
          var,
          classe,
          valor
        FROM medicoes
        WHERE {where_sql}
        ORDER BY timestamp ASC;
    """
    return con.execute(q).df()


def _normalize_long_columns(df_long: pd.DataFrame) -> pd.DataFrame:
    canonical_names = {
        "timestamp": "timestamp",
        "equip_id": "equip_id",
        "var": "var",
        "classe": "classe",
        "valor": "valor",
        "se": "SE",
        "bay": "BAY",
        "terminal": "TERMINAL",
    }
    lower_to_actual = {str(col).lower(): col for col in df_long.columns}
    rename_map = {}
    for lowered, canonical in canonical_names.items():
        actual = lower_to_actual.get(lowered)
        if actual and actual != canonical:
            rename_map[actual] = canonical
    if not rename_map:
        return df_long
    return df_long.rename(columns=rename_map)


def aggregate_long(
    df_long: pd.DataFrame,
    agg: str = "max",
    time_floor: Optional[str] = None,
) -> pd.DataFrame:
    if df_long.empty:
        return pd.DataFrame(columns=["_TS", "_KEY", "_VAL", "_EQUIP", "_VAR", "_CLASSE"])

    agg = agg.lower().strip()
    if agg not in {"max", "last"}:
        raise ValueError("agg deve ser 'max' ou 'last'")

    df_long = _normalize_long_columns(df_long)

    # Seleciona apenas colunas necessarias para reduzir RAM.
    base_cols = ["timestamp", "equip_id", "var", "classe", "valor"]
    opt_cols = [c for c in ["SE", "BAY", "TERMINAL"] if c in df_long.columns]
    df = df_long[base_cols + opt_cols].copy()
    ts = pd.to_datetime(df["timestamp"], errors="coerce")
    # Normaliza timezone e remove microssegundos para garantir chaves estaveis.
    if getattr(ts.dt, "tz", None) is not None:
        ts = ts.dt.tz_convert(None)
    ts = ts.dt.floor("s")
    df["timestamp"] = ts
    df = df.dropna(subset=["timestamp", "equip_id", "var"])

    if time_floor:
        df["timestamp"] = df["timestamp"].dt.floor(time_floor)

    df["_EQUIP"] = df["equip_id"].astype(str)
    df["_VAR"] = df["var"].astype(str)
    df["_CLASSE"] = df["classe"].fillna("").astype(str)
    df["_SE"] = df["SE"].astype(str) if "SE" in df.columns else ""
    df["_BAY"] = df["BAY"].astype(str) if "BAY" in df.columns else ""
    df["_TERMINAL"] = df["TERMINAL"].astype(str) if "TERMINAL" in df.columns else ""
    df["_TS"] = df["timestamp"]
    df["_KEY"] = (
        df["_SE"].fillna("").astype(str)
        + "|"
        + df["_BAY"].fillna("").astype(str)
        + "|"
        + df["_EQUIP"].fillna("").astype(str)
        + "|"
        + df["_TERMINAL"].fillna("").astype(str)
        + "|"
        + df["_VAR"].fillna("").astype(str)
    )
    df["_VAL"] = normalize_measurement_series(df["valor"], ndigits=1)
    df = df.dropna(subset=["_VAL"])

    if df.empty:
        return pd.DataFrame(columns=["_TS", "_KEY", "_VAL", "_EQUIP", "_VAR", "_CLASSE"])

    if agg == "max":
        out = df.groupby(
            ["_TS", "_KEY", "_SE", "_BAY", "_TERMINAL", "_EQUIP", "_VAR", "_CLASSE"],
            as_index=False,
        )["_VAL"].max()
    else:
        df = df.sort_values(["_TS"])
        out = df.groupby(["_TS", "_KEY"], as_index=False).last()
        out = out[["_TS", "_KEY", "_SE", "_BAY", "_TERMINAL", "_EQUIP", "_VAR", "_CLASSE", "_VAL"]]
    out["_VAL"] = normalize_measurement_series(out["_VAL"], ndigits=1)

    return out


def _estimate_suggestions_for_empty(
    ov: Dict[str, str],
    equips: List[str],
    t0: str,
    t1: str,
    vars_: Optional[List[str]],
) -> str:
    _ = equips
    tips = [
        "Verifique se os equipamentos estao corretos (--equip).",
        f"Verifique se o intervalo ({t0} ate {t1}) esta dentro do range do lake ({ov['ts_min']} ate {ov['ts_max']}).",
    ]
    if vars_:
        tips.append("Verifique se as variaveis informadas em --var existem para esses equipamentos.")
    return " ".join(tips)


def construir_bi_excel_multi_equip_multi_var_long(
    long_df: pd.DataFrame,
    xlsx_out: str,
    report_meta: Dict[str, str],
    equip_slots: int = 8,
    var_slots: int = 6,
    max_timestamps: int = 300_000,
    selected_pairs: Optional[List[Tuple[str, str]]] = None,
    split_timestamp_columns: bool = False,
) -> Tuple[str, List[str]]:
    if long_df.empty:
        raise ValueError("Recorte vazio: nao ha dados para gerar dashboard.")

    warnings: List[str] = []
    df = long_df.copy()
    df["_TS"] = pd.to_datetime(df["_TS"], errors="coerce")
    df = df.dropna(subset=["_TS", "_KEY"])
    if df.empty:
        raise ValueError("Recorte vazio apos normalizacao do timestamp.")

    equips = sorted(df["_EQUIP"].dropna().astype(str).unique().tolist())
    vars_detectadas = sorted(df["_VAR"].dropna().astype(str).unique().tolist())

    var_class = (
        df.dropna(subset=["_VAR"])
        .groupby("_VAR")["_CLASSE"]
        .agg(lambda x: x.dropna().astype(str).iloc[0] if len(x.dropna()) else "")
        .to_dict()
    )

    ts_unique = pd.to_datetime(df["_TS"], errors="coerce").dropna().dt.floor("s").drop_duplicates().sort_values()
    ts_before = len(ts_unique)
    if ts_before > max_timestamps:
        ts_unique = ts_unique.iloc[:max_timestamps]
        keep = set(ts_unique.astype("datetime64[ns]").tolist())
        df = df[df["_TS"].isin(keep)]
        warnings.append(
            f"Timestamps truncados: {ts_before} -> {max_timestamps} (max_timestamps). "
            "Reduza intervalo ou aumente timefloor."
        )

    max_data_rows = EXCEL_MAX_ROWS - 1
    if len(df) > max_data_rows:
        warnings.append(
            f"DADOS_AGG excede limite do Excel ({len(df)} linhas > {max_data_rows}). "
            "Truncando saida. Reduza intervalo/vars ou aumente timefloor."
        )
        df = df.iloc[:max_data_rows].copy()

    df["_TS"] = pd.to_datetime(df["_TS"], errors="coerce").dt.floor("s")
    df = df.dropna(subset=["_TS"]).sort_values(["_TS", "_KEY"]).copy()
    ts_unique = df["_TS"].drop_duplicates().sort_values()
    n_points = len(ts_unique)
    time_gaps, time_audit = build_temporal_gap_audit(ts_unique)
    if int(time_audit.get("gap_count", 0) or 0) > 0:
        warnings.append(
            "Auditoria temporal detectou "
            f"{int(time_audit['gap_count'])} lacuna(s) no recorte; "
            f"primeira apos {time_audit.get('first_gap_after', '') or '(inicio)'}."
        )

    wb_out = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(color="FFFFFF", bold=True)

    ws_agg = wb_out.active
    ws_agg.title = "DADOS_AGG"

    headers = ["TIMESTAMP"]
    if split_timestamp_columns:
        headers.extend([TIMESTAMP_SPLIT_DATE_COLUMN, TIMESTAMP_SPLIT_TIME_COLUMN])
    headers.extend(["KEY", "VAL", "EQUIP", "VAR", "CLASSE", "TSKEY"])
    for j, h in enumerate(headers, start=1):
        c = ws_agg.cell(1, j, h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    df["_TS_TXT"] = df["_TS"].dt.strftime(TS_PLOT_FMT)
    df["TSKEY"] = df["_TS_TXT"] + "|" + df["_KEY"]

    # Streaming write: evita materializar lista gigante em memoria.
    for i, row in enumerate(
        df[["_TS", "_KEY", "_VAL", "_EQUIP", "_VAR", "_CLASSE", "TSKEY"]].itertuples(index=False, name=None),
        start=2,
    ):
        ts_value = pd.Timestamp(row[0]) if pd.notna(row[0]) else None
        current_col = 1
        ts_cell = ws_agg.cell(i, current_col, ts_value.to_pydatetime() if ts_value is not None else None)
        ts_cell.number_format = "yyyy-mm-dd hh:mm:ss"
        current_col += 1
        if split_timestamp_columns:
            if ts_value is not None:
                date_value = ts_value.to_pydatetime().date()
                time_value = ts_value.to_pydatetime().time().replace(microsecond=0)
            else:
                date_value = None
                time_value = None
            date_cell = ws_agg.cell(i, current_col, date_value)
            date_cell.number_format = TIMESTAMP_SPLIT_EXCEL_DATE_NUMBER_FORMAT
            current_col += 1
            time_cell = ws_agg.cell(i, current_col, time_value)
            time_cell.number_format = TIMESTAMP_SPLIT_EXCEL_TIME_NUMBER_FORMAT
            current_col += 1
        for value_idx, val in enumerate(row[1:], start=1):
            if value_idx == 2:
                val = round_half_up_float(val, ndigits=1)
            cell = ws_agg.cell(i, current_col, val)
            if value_idx == 2:
                cell.number_format = "0.0"
            current_col += 1

    ws_agg.freeze_panes = "A2"
    ws_agg.column_dimensions["A"].width = 20
    if split_timestamp_columns:
        ws_agg.column_dimensions["B"].width = 12
        ws_agg.column_dimensions["C"].width = 12
        ws_agg.column_dimensions["D"].width = 50
        ws_agg.column_dimensions["E"].width = 14
        ws_agg.column_dimensions["I"].width = 70
    else:
        ws_agg.column_dimensions["B"].width = 50
        ws_agg.column_dimensions["C"].width = 14
        ws_agg.column_dimensions["G"].width = 70
    ws_agg.sheet_state = "hidden"

    ws_cfg = wb_out.create_sheet("CONFIG")
    ws_cfg["A1"] = f"CONFIG - Parametros do dashboard ({ENV_LABEL})"
    ws_cfg["A1"].font = Font(bold=True, size=16)

    ws_cfg["A3"] = "Classe (POT/COR/TEN):"
    ws_cfg["A4"] = "Variaveis (VAR1..):"
    ws_cfg["A5"] = "Equipamentos (EQUIP1..):"
    ws_cfg["A2"] = "MAX_SERIES_POR_GRAFICO:"
    for a in ["A3", "A4", "A5"]:
        ws_cfg[a].font = Font(bold=True)
    ws_cfg["A2"].font = Font(bold=True)

    for i in range(var_slots):
        ws_cfg[f"A{7+i}"] = f"VAR {i+1}:"
        ws_cfg[f"A{7+i}"].font = Font(bold=True)

    equip_start_row = 7 + var_slots + 2
    ws_cfg[f"A{equip_start_row-1}"] = "-"
    for i in range(equip_slots):
        ws_cfg[f"A{equip_start_row+i}"] = f"EQUIP {i+1}:"
        ws_cfg[f"A{equip_start_row+i}"].font = Font(bold=True)

    ws_cfg["D1"] = "EQUIPS"
    for i, e in enumerate(equips, start=2):
        ws_cfg.cell(i, 4, e)

    vars_pot = [v for v in vars_detectadas if var_class.get(v) == "POT"]
    vars_cor = [v for v in vars_detectadas if var_class.get(v) == "COR"]
    vars_ten = [v for v in vars_detectadas if var_class.get(v) == "TEN"]

    ws_cfg["E1"] = "VAR_POT"
    for i, v in enumerate(vars_pot, start=2):
        ws_cfg.cell(i, 5, v)
    ws_cfg["F1"] = "VAR_COR"
    for i, v in enumerate(vars_cor, start=2):
        ws_cfg.cell(i, 6, v)
    ws_cfg["G1"] = "VAR_TEN"
    for i, v in enumerate(vars_ten, start=2):
        ws_cfg.cell(i, 7, v)

    add_named_range(wb_out, "EQUIPS", "CONFIG", "D", len(equips))
    add_named_range(wb_out, "VAR_POT", "CONFIG", "E", len(vars_pot))
    add_named_range(wb_out, "VAR_COR", "CONFIG", "F", len(vars_cor))
    add_named_range(wb_out, "VAR_TEN", "CONFIG", "G", len(vars_ten))

    ws_cfg["B3"] = "COR" if vars_cor else ("TEN" if vars_ten else "POT")
    max_plot_cap = 24
    ws_cfg["B2"] = 6
    default_vars = vars_cor if ws_cfg["B3"].value == "COR" else (vars_ten if ws_cfg["B3"].value == "TEN" else vars_pot)
    for i in range(var_slots):
        ws_cfg[f"B{7+i}"] = default_vars[i] if i < min(len(default_vars), 3) else ""

    for i in range(equip_slots):
        ws_cfg[f"B{equip_start_row+i}"] = equips[i] if i < min(len(equips), 3) else ""

    dv_class = DataValidation(type="list", formula1='"POT,COR,TEN"', allow_blank=False)
    ws_cfg.add_data_validation(dv_class)
    dv_class.add(ws_cfg["B3"])

    max_choices = ",".join(str(i) for i in range(1, max_plot_cap + 1))
    dv_max_series = DataValidation(type="list", formula1=f'"{max_choices}"', allow_blank=False)
    ws_cfg.add_data_validation(dv_max_series)
    dv_max_series.add(ws_cfg["B2"])

    dv_var = DataValidation(type="list", formula1='=INDIRECT("VAR_"&$B$3)', allow_blank=True)
    ws_cfg.add_data_validation(dv_var)
    for i in range(var_slots):
        dv_var.add(ws_cfg[f"B{7+i}"])

    dv_equip = DataValidation(type="list", formula1="=EQUIPS", allow_blank=True)
    ws_cfg.add_data_validation(dv_equip)
    for i in range(equip_slots):
        dv_equip.add(ws_cfg[f"B{equip_start_row+i}"])

    ws_cfg.column_dimensions["A"].width = 26
    ws_cfg.column_dimensions["B"].width = 60
    ws_cfg.freeze_panes = "A7"
    for col in ["D", "E", "F", "G"]:
        ws_cfg.column_dimensions[col].hidden = True

    ws_cfg["A6"] = (
        "Dica: em exports do app, todas as series validas entram na aba tecnica oculta "
        "PLOT_DATA; o agrupamento divide os graficos para preservar leitura."
    )
    ws_cfg["A6"].font = Font(size=10, color="666666")

    ws_view = wb_out.create_sheet("VIEW_MX")
    ws_view["A1"] = "VIEW_MX - Tabela dinamica das series exportadas"
    ws_view["A1"].font = Font(bold=True, size=16)
    ws_view["A2"] = "Colunas = series com identidade completa (SE/BAY/EQUIP/TERMINAL/VAR)."
    ws_view["A2"].font = Font(size=10, color="666666")

    base_row = 4
    series_start_col = 2
    if split_timestamp_columns:
        ws_view.cell(base_row, 1, TIMESTAMP_SPLIT_DATE_COLUMN)
        ws_view.cell(base_row, 2, TIMESTAMP_SPLIT_TIME_COLUMN)
        series_start_col = 3
    else:
        ws_view.cell(base_row, 1, "TIMESTAMP_TEXT")
    ws_view.cell(base_row + 1, 1, "EQUIP")
    ws_view.cell(base_row + 2, 1, "VAR")
    ws_view.cell(base_row + 3, 1, "KEY")

    # Preenche VIEW_MX por pivot em Python (sem lookup no Excel).
    wide = (
        df.pivot_table(index="_TS", columns="_KEY", values="_VAL", aggfunc="last")
        .sort_index()
        .reset_index()
    )
    wide["_TS_TXT"] = pd.to_datetime(wide["_TS"], errors="coerce").dt.strftime(TS_PLOT_FMT)
    wide_columns = [c for c in wide.columns if c != "_TS"]
    wide_columns = [c for c in wide_columns if c != "_TS_TXT"]
    plot_specs = resolve_dashboard_series_specs(wide_columns, selected_pairs=selected_pairs)
    if not plot_specs:
        plot_specs = resolve_dashboard_series_specs(wide_columns, selected_pairs=None)

    cc = series_start_col
    for raw_key in wide_columns:
        spec = parse_dashboard_series_key(raw_key)
        ws_view.cell(base_row, cc, build_dashboard_series_label(spec))
        ws_view.cell(base_row + 1, cc, spec.equip)
        ws_view.cell(base_row + 2, cc, spec.var)
        ws_view.cell(base_row + 3, cc, spec.key)
        cc += 1

    for r in (base_row, base_row + 1, base_row + 2, base_row + 3):
        for colx in range(1, cc):
            cell = ws_view.cell(r, colx)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_view.freeze_panes = f"{openpyxl.utils.get_column_letter(series_start_col)}{base_row+4}"
    ws_view.column_dimensions["A"].width = 22

    series_keys = [c for c in wide.columns if c not in {"_TS", "_TS_TXT"}]
    view_frame = wide[["_TS", "_TS_TXT"] + series_keys]
    for i, row in enumerate(view_frame.itertuples(index=False, name=None), start=base_row + 4):
        ts_value = pd.Timestamp(row[0]) if pd.notna(row[0]) else None
        if split_timestamp_columns:
            if ts_value is not None:
                date_value = ts_value.to_pydatetime().date()
                time_value = ts_value.to_pydatetime().time().replace(microsecond=0)
            else:
                date_value = None
                time_value = None
            date_cell = ws_view.cell(i, 1, date_value)
            date_cell.number_format = TIMESTAMP_SPLIT_EXCEL_DATE_NUMBER_FORMAT
            time_cell = ws_view.cell(i, 2, time_value)
            time_cell.number_format = TIMESTAMP_SPLIT_EXCEL_TIME_NUMBER_FORMAT
        else:
            ws_view.cell(i, 1, row[1])
        for j, value in enumerate(row[2:], start=series_start_col):
            ws_view.cell(i, j, value)

    for colv in range(2, cc):
        ws_view.column_dimensions[openpyxl.utils.get_column_letter(colv)].width = 16
    if split_timestamp_columns:
        ws_view.column_dimensions["A"].width = 12
        ws_view.column_dimensions["B"].width = 12

    # Layout mais legivel para cabecalhos
    ws_view.row_dimensions[base_row].height = 24
    ws_view.row_dimensions[base_row + 1].height = 22
    ws_view.row_dimensions[base_row + 2].height = 22
    ws_view.row_dimensions[base_row + 3].height = 24
    if not split_timestamp_columns:
        ws_view.column_dimensions["A"].width = 22

    ws_plot = wb_out.create_sheet(PLOT_DATA_SHEET)
    ws_plot["A1"] = "TIMESTAMP_TEXT"
    ws_plot["A1"].fill = head_fill
    ws_plot["A1"].font = head_font
    ws_plot["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_plot.column_dimensions["A"].width = 22
    for col_idx, spec in enumerate(plot_specs, start=2):
        header = ws_plot.cell(1, col_idx, build_dashboard_series_label(spec))
        header.fill = head_fill
        header.font = head_font
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_plot.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24

    plot_frame = wide[["_TS_TXT"] + [spec.key for spec in plot_specs]].copy()
    for row_idx, row in enumerate(plot_frame.itertuples(index=False, name=None), start=2):
        ws_plot.cell(row_idx, 1, row[0])
        for col_idx, value in enumerate(row[1:], start=2):
            if pd.isna(value):
                ws_plot.cell(row_idx, col_idx, None)
            else:
                ws_plot.cell(row_idx, col_idx, round_half_up_float(float(value), ndigits=1))
                ws_plot.cell(row_idx, col_idx).number_format = "0.0"
    ws_plot.freeze_panes = "A2"
    ws_plot.sheet_state = "hidden"

    ws_dash = wb_out.create_sheet("DASHBOARD")
    ws_dash.sheet_view.showGridLines = False
    series_per_chart = max(1, min(int(ws_cfg["B2"].value or 6), max_plot_cap))
    chart_groups = build_dashboard_chart_groups(plot_specs, max_series_per_chart=series_per_chart)
    total_series = len(plot_specs)
    audit_entries = build_dashboard_series_audit(plot_specs, chart_groups)
    plotted_series_total = sum(1 for entry in audit_entries if entry.plotted)
    if plotted_series_total != total_series:
        missing_labels = [entry.label for entry in audit_entries if not entry.plotted]
        raise RuntimeError(
            "Integridade do DASHBOARD violada: "
            f"{plotted_series_total}/{total_series} series foram alocadas em graficos. "
            f"Series ausentes: {missing_labels[:8]}"
        )

    summary_fill = PatternFill("solid", fgColor="D9EAF7")

    def _merge_dashboard(range_ref: str, value: str, *, bold: bool = False, size: int = 11, color: Optional[str] = None, fill: Optional[PatternFill] = None) -> None:
        ws_dash.merge_cells(range_ref)
        cell = ws_dash[range_ref.split(":")[0]]
        cell.value = neutralize_formula_value(value)
        cell.font = Font(bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if fill is not None:
            cell.fill = fill

    for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"):
        ws_dash.column_dimensions[col].width = 12
    ws_dash.column_dimensions["A"].width = 4
    ws_dash.column_dimensions["N"].width = 4

    _merge_dashboard("B1:M1", "DASHBOARD - Relatorio personalizado (Andy's Lake)", bold=True, size=18)
    _merge_dashboard(
        "B2:M2",
        "O dashboard plota todas as series exportadas e adapta automaticamente legenda, espacamento e escala do eixo Y para preservar legibilidade tecnica.",
        size=11,
    )

    ws_dash["B4"] = "Classe"
    ws_dash["D4"] = "=CONFIG!B3"
    ws_dash["F4"] = "Series exportadas"
    ws_dash["H4"] = str(total_series)
    ws_dash["J4"] = "Series plotadas"
    ws_dash["L4"] = str(plotted_series_total)
    for cell_ref in ("B4", "F4", "J4"):
        ws_dash[cell_ref].font = Font(bold=True)
        ws_dash[cell_ref].fill = summary_fill
    ws_dash["B4"].fill = summary_fill
    ws_dash["F4"].fill = summary_fill
    ws_dash["J4"].fill = summary_fill
    ws_dash["D4"].font = Font(bold=True)
    ws_dash["H4"].font = Font(bold=True)
    ws_dash["L4"].font = Font(bold=True)

    _merge_dashboard("B5:C5", "Organizacao", bold=True, fill=summary_fill)
    _merge_dashboard("D5:M5", f"{len(chart_groups)} grafico(s) | agrupamento por VAR | ate {series_per_chart} serie(s)/grafico")
    _merge_dashboard("B6:C6", "Intervalo", bold=True, fill=summary_fill)
    _merge_dashboard("D6:G6", report_meta.get("period_label", f"{report_meta.get('t0', '')} -> {report_meta.get('t1', '')}"))
    _merge_dashboard("H6:I6", "Agg/Timefloor", bold=True, fill=summary_fill)
    _merge_dashboard("J6:M6", f"{report_meta.get('agg', '')} / {report_meta.get('time_floor', '(none)')}")
    _merge_dashboard("B7:C7", "Entradas", bold=True, fill=summary_fill)
    _merge_dashboard("D7:G7", report_meta.get("equips", ""))
    _merge_dashboard("H7:I7", "Variaveis", bold=True, fill=summary_fill)
    _merge_dashboard("J7:M7", report_meta.get("vars", "(todas)"))
    _merge_dashboard("B8:C8", "Legenda", bold=True, fill=summary_fill)
    _merge_dashboard("D8:G8", "Direita ate 4 series; abaixo acima disso")
    _merge_dashboard("H8:I8", "Auditoria", bold=True, fill=summary_fill)
    _merge_dashboard("J8:M8", f"PLOT_DATA(oculta)={total_series} | plotadas={plotted_series_total} | perda=0")
    _merge_dashboard("B9:C9", "Range real", bold=True, fill=summary_fill)
    _merge_dashboard("D9:H9", f"{report_meta.get('ts_min', '')} -> {report_meta.get('ts_max', '')}")
    _merge_dashboard("I9:J9", "Gerado em", bold=True, fill=summary_fill)
    _merge_dashboard("K9:M9", report_meta.get("generated_at", ""))
    _merge_dashboard(
        "B10:M10",
        "A base dos graficos fica na aba tecnica oculta PLOT_DATA. VIEW_MX, TIME_AUDIT e PLOT_AUDIT permanecem disponiveis para inspecao tecnica e conciliacao das series.",
        size=10,
        color="666666",
    )

    for row_idx in range(1, 11):
        ws_dash.row_dimensions[row_idx].height = 22

    chart_plans = build_dashboard_chart_layouts(chart_groups, plot_frame)
    ws_dash.freeze_panes = f"B{DASHBOARD_CHART_START_ROW}"

    if chart_plans:
        cats_ref = Reference(ws_plot, min_col=1, min_row=2, max_row=1 + n_points)
        key_to_column = {spec.key: 2 + idx for idx, spec in enumerate(plot_specs)}
        for row_idx in range(DASHBOARD_CHART_START_ROW, chart_plans[-1].anchor_to_row + 1):
            ws_dash.row_dimensions[row_idx].height = DASHBOARD_CHART_ROW_HEIGHT_PT

        for group, plan in zip(chart_groups, chart_plans):
            min_chart_col = min(key_to_column[item.key] for item in group.series)
            max_chart_col = max(key_to_column[item.key] for item in group.series)
            chart = LineChart()
            chart.title = group.title
            chart.y_axis.title = "Valor real (com sinal)"
            chart.x_axis.title = "Data/Hora"
            chart.legend.position = plan.legend_position
            chart.legend.overlay = False
            chart.style = 2
            chart.plotVisOnly = False
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.crosses = plan.axis.crosses_mode
            chart.y_axis.numFmt = plan.axis.num_format
            chart.y_axis.majorGridlines = ChartLines()
            if plan.axis.min_value is not None:
                chart.y_axis.scaling.min = plan.axis.min_value
            if plan.axis.max_value is not None:
                chart.y_axis.scaling.max = plan.axis.max_value
            chart.add_data(
                Reference(
                    ws_plot,
                    min_col=min_chart_col,
                    min_row=1,
                    max_col=max_chart_col,
                    max_row=1 + n_points,
                ),
                titles_from_data=True,
            )
            chart.set_categories(cats_ref)
            for s in chart.series:
                s.marker.symbol = "circle"
                s.marker.size = 4 if len(group.series) > 4 else 5
                s.smooth = False
                s.graphicalProperties.line.width = 9000 if len(group.series) > 4 else 10000
            chart.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=plan.anchor_from_col - 1, row=plan.anchor_from_row - 1, colOff=0, rowOff=0),
                to=AnchorMarker(col=plan.anchor_to_col - 1, row=plan.anchor_to_row - 1, colOff=0, rowOff=0),
            )
            ws_dash.add_chart(chart)

        if len(chart_groups) > 1:
            warnings.append(
                f"DASHBOARD dividido em {len(chart_groups)} graficos para manter legibilidade "
                f"({series_per_chart} serie(s) por grafico, agrupadas por VAR)."
            )
    else:
        warnings.append("Sem series para plotar no DASHBOARD (chaves de variavel vazias).")

    ws_audit = wb_out.create_sheet("PLOT_AUDIT")
    audit_headers = ["SERIE_KEY", "SERIE_LABEL", "PLOT_DATA_COL", "PLOTTED", "CHART_GROUP"]
    for col_idx, header in enumerate(audit_headers, start=1):
        cell = ws_audit.cell(1, col_idx, header)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, entry in enumerate(audit_entries, start=2):
        ws_audit.cell(row_idx, 1, entry.key)
        ws_audit.cell(row_idx, 2, entry.label)
        ws_audit.cell(row_idx, 3, entry.plot_column)
        ws_audit.cell(row_idx, 4, "YES" if entry.plotted else "NO")
        ws_audit.cell(row_idx, 5, entry.chart_title or "-")
    ws_audit.freeze_panes = "A2"
    ws_audit.column_dimensions["A"].width = 42
    ws_audit.column_dimensions["B"].width = 48
    ws_audit.column_dimensions["C"].width = 14
    ws_audit.column_dimensions["D"].width = 12
    ws_audit.column_dimensions["E"].width = 36

    ws_time = wb_out.create_sheet(TIME_AUDIT_SHEET)
    time_headers = ["PREV_TIMESTAMP", "CUR_TIMESTAMP", "DELTA_MINUTES", "EXPECTED_MINUTES", "MISSING_INTERVALS", "GAP_KIND"]
    for col_idx, header in enumerate(time_headers, start=1):
        cell = ws_time.cell(1, col_idx, header)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if time_gaps.empty:
        ws_time.cell(2, 1, None)
        ws_time.cell(2, 2, None)
        ws_time.cell(2, 3, 0)
        ws_time.cell(2, 4, float(time_audit.get("expected_minutes", 0.0) or 0.0))
        ws_time.cell(2, 5, 0)
        ws_time.cell(2, 6, "NO_GAPS_DETECTED")
    else:
        for row_idx, row in enumerate(time_gaps.itertuples(index=False), start=2):
            ws_time.cell(row_idx, 1, row.prev_ts.strftime(TS_FMT) if pd.notna(row.prev_ts) else "")
            ws_time.cell(row_idx, 2, row.curr_ts.strftime(TS_FMT) if pd.notna(row.curr_ts) else "")
            ws_time.cell(row_idx, 3, float(row.delta_minutes))
            ws_time.cell(row_idx, 4, float(row.expected_minutes))
            ws_time.cell(row_idx, 5, int(row.missing_intervals))
            ws_time.cell(row_idx, 6, str(row.gap_kind))
    ws_time.freeze_panes = "A2"
    ws_time.column_dimensions["A"].width = 22
    ws_time.column_dimensions["B"].width = 22
    ws_time.column_dimensions["C"].width = 16
    ws_time.column_dimensions["D"].width = 18
    ws_time.column_dimensions["E"].width = 18
    ws_time.column_dimensions["F"].width = 18

    wb_out.active = wb_out.index(ws_dash)

    _ensure_dir(os.path.dirname(xlsx_out))
    try:
        wb_out.save(xlsx_out)
    except PermissionError as e:
        raise PermissionError(
            f"Falha ao salvar XLSX em {xlsx_out}. O arquivo pode estar aberto no Excel ou sem permissao."
        ) from e
    except OSError as e:
        raise OSError(f"Falha ao salvar XLSX em {xlsx_out}: {e}") from e
    return xlsx_out, warnings


def construir_bi_excel_v2(
    long_df: pd.DataFrame,
    xlsx_out: str,
    report_meta: Dict[str, str],
    *,
    template_path: Optional[str] = None,
    patamar_result: Optional[PatamarResult] = None,
    patamar_write_spec: Optional[PatamarWriteSpec] = None,
    **kwargs: object,
) -> Tuple[str, List[str]]:
    out_file, warnings = construir_bi_excel_multi_equip_multi_var_long(
        long_df,
        xlsx_out,
        report_meta,
        equip_slots=int(kwargs.get("equip_slots", 8) or 8),
        var_slots=int(kwargs.get("var_slots", 6) or 6),
        max_timestamps=int(kwargs.get("max_timestamps", 300_000) or 300_000),
        selected_pairs=kwargs.get("selected_pairs"),  # type: ignore[arg-type]
        split_timestamp_columns=bool(kwargs.get("split_timestamp_columns", False)),
    )
    if patamar_result is not None:
        _write_patamar_summary(
            out_file,
            patamar_result,
            patamar_write_spec or PatamarWriteSpec(),
            template_path=template_path,
        )
    return out_file, warnings


def _write_patamar_summary(
    workbook_path: str,
    patamar_result: PatamarResult,
    spec: PatamarWriteSpec,
    *,
    template_path: Optional[str] = None,
) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    try:
        if spec.summary_sheet not in wb.sheetnames:
            wb.create_sheet(spec.summary_sheet)
        ws = wb[spec.summary_sheet]
        for idx, shift in enumerate(patamar_result.shifts, start=spec.first_row):
            ws[f"{spec.p_column}{idx}"] = str(shift.p_text)
            ws[f"{spec.q_column}{idx}"] = str(shift.q_text)
        if template_path and os.path.exists(template_path):
            ws["A1"] = ws["A1"].value or "PATAMAR - Resumo"
        wb.save(workbook_path)
    finally:
        wb.close()


def run_report(
    work_root: str,
    db_path: str,
    equips: List[str],
    t0: str,
    t1: str,
    vars_: Optional[List[str]],
    agg: str,
    time_floor: Optional[str],
    equip_slots: int,
    var_slots: int,
    max_timestamps: int,
    out_dir: str,
    out_name: Optional[str] = None,
) -> str:
    _ = work_root
    log("LAKE", f"Ambiente: {ENV_LABEL}")
    log("LAKE", f"DB: {db_path}")

    _ensure_dir(out_dir)
    validate_ts_or_raise(t0, "--from")
    validate_ts_or_raise(t1, "--to")
    if t0 > t1:
        raise ValueError("Intervalo invalido: --from deve ser <= --to.")

    con = connect_duckdb(db_path)
    try:
        ensure_medicoes_exists(con)
        ov = lake_overview(con)
        log("LAKE", f"overview={ov}")

        where_sql = build_where_sql(equips=equips, t0=t0, t1=t1, vars_=vars_)
        n_raw = count_recorte(con, where_sql)
        log("QUERY", f"linhas estimadas no recorte: {n_raw}")

        effective_time_floor = time_floor
        if n_raw >= AUTO_TIMEFLOOR_ROWS and not time_floor:
            effective_time_floor = "15min"
            log(
                "WARN",
                "Recorte muito grande. Aplicando timefloor automatico='15min'. "
                "Opcionalmente ajuste --timefloor e reduza --var.",
            )
        elif n_raw >= LARGE_ROWS_WARN:
            log(
                "WARN",
                "Recorte grande. Considere --timefloor (ex: 15min) e reduzir variaveis para diminuir RAM/tempo.",
            )

        if n_raw == 0:
            raise ValueError(
                "Recorte vazio na consulta ao lake. "
                + _estimate_suggestions_for_empty(ov=ov, equips=equips, t0=t0, t1=t1, vars_=vars_)
            )

        log("QUERY", "Extraindo recorte LONG...")
        df_long = fetch_recorte_long(con, where_sql=where_sql)
        log("QUERY", f"linhas LONG brutas: {len(df_long)}")
    finally:
        con.close()

    log("AGG", f"Agregando com agg={agg}, timefloor={effective_time_floor or '(none)'}")
    df_agg = aggregate_long(df_long, agg=agg, time_floor=effective_time_floor)
    log("AGG", f"linhas LONG agregadas: {len(df_agg)}")
    if df_agg.empty:
        raise ValueError(
            "Recorte ficou vazio apos agregacao. "
            "Possiveis causas: equip errado, intervalo fora do range, var inexistente ou valores nao numericos."
        )

    if not out_name:
        now = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_e = re.sub(r"[^A-Za-z0-9_-]+", "_", "_".join(equips[:3]))
        out_name = f"RELATORIO_ANDYS_{safe_e}_{now}.xlsx"
    xlsx_out = os.path.join(out_dir, out_name)

    meta = {
        "equips": ", ".join(equips),
        "t0": t0,
        "t1": t1,
        "vars": ", ".join(vars_ or []),
        "agg": agg,
        "time_floor": effective_time_floor or "(none)",
        "ts_min": str(df_agg["_TS"].min()) if not df_agg.empty else "",
        "ts_max": str(df_agg["_TS"].max()) if not df_agg.empty else "",
        "n_ts_unique": str(df_agg["_TS"].nunique()) if not df_agg.empty else "0",
        "generated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    log("XLSX", "Gerando dashboard...")
    out, warns = construir_bi_excel_multi_equip_multi_var_long(
        long_df=df_agg,
        xlsx_out=xlsx_out,
        report_meta=meta,
        equip_slots=equip_slots,
        var_slots=var_slots,
        max_timestamps=max_timestamps,
    )
    for w in warns:
        log("WARN", w)
    log("XLSX", f"Dashboard gerado: {out}")
    return out


def run_selftest(db_path: str, out_dir: str, work_root: str) -> str:
    log("LAKE", "Iniciando --selftest")
    con = connect_duckdb(db_path)
    try:
        ensure_medicoes_exists(con)
        ov = lake_overview(con)
        log("LAKE", f"overview={ov}")
        row = con.execute(
            """
            SELECT equip_id, MIN(timestamp) AS ts0
            FROM medicoes
            GROUP BY equip_id
            ORDER BY COUNT(*) DESC
            LIMIT 1;
            """
        ).fetchone()
        if row is None:
            raise RuntimeError("Selftest falhou: lake sem dados em medicoes.")
        equip = str(row[0])
        ts0 = pd.to_datetime(row[1]).to_pydatetime().replace(minute=0, second=0, microsecond=0)
        t0 = ts0.strftime(TS_FMT)
        t1 = (ts0 + dt.timedelta(hours=1)).strftime(TS_FMT)
        vars_rows = con.execute(
            f"""
            SELECT DISTINCT var
            FROM medicoes
            WHERE equip_id = {sql_quote(equip)}
            ORDER BY var;
            """
        ).fetchall()
        vars_all = [str(v[0]) for v in vars_rows]
        preferred = [v for v in ["IA", "IB", "IC"] if v in vars_all]
        vars_sel = preferred if preferred else vars_all[:3]
        if not vars_sel:
            raise RuntimeError("Selftest falhou: nao encontrou variaveis para o equipamento.")
        log("QUERY", f"selftest equip={equip} t0={t0} t1={t1}")
    finally:
        con.close()

    out_name = f"SELFTEST_ANDYS_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out = run_report(
        work_root=work_root,
        db_path=db_path,
        equips=[equip],
        t0=t0,
        t1=t1,
        vars_=vars_sel,
        agg="max",
        time_floor="15min",
        equip_slots=1,
        var_slots=3,
        max_timestamps=2_000,
        out_dir=out_dir,
        out_name=out_name,
    )
    if not os.path.exists(out):
        raise RuntimeError(f"Selftest falhou: arquivo nao criado ({out}).")

    wb_chk = openpyxl.load_workbook(out, data_only=False, read_only=False)
    try:
        has_xlookup = False
        ws_view = wb_chk["VIEW_MX"] if "VIEW_MX" in wb_chk.sheetnames else None
        non_empty_numeric = 0
        for ws in wb_chk.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        u = v.upper()
                        if "XLOOKUP(" in u:
                            has_xlookup = True
                    if ws_view is not None and ws.title == "VIEW_MX" and isinstance(v, (int, float)):
                        non_empty_numeric += 1
        if has_xlookup:
            raise RuntimeError("Selftest falhou: formula XLOOKUP encontrada no XLSX.")
        if non_empty_numeric == 0:
            raise RuntimeError("Selftest falhou: VIEW_MX sem valores numericos preenchidos.")
        ws_dash = wb_chk["DASHBOARD"] if "DASHBOARD" in wb_chk.sheetnames else None
        if ws_dash is None:
            raise RuntimeError("Selftest falhou: aba DASHBOARD ausente.")
        if any(cell.value == PLOT_DATA_SHEET for cell in ws_dash["A"]):
            raise RuntimeError("Selftest falhou: bloco tabular PLOT_DATA permaneceu visivel no DASHBOARD.")
        ws_plot = wb_chk[PLOT_DATA_SHEET] if PLOT_DATA_SHEET in wb_chk.sheetnames else None
        if ws_plot is None:
            raise RuntimeError("Selftest falhou: aba PLOT_DATA ausente.")
        if ws_plot.sheet_state not in {"hidden", "veryHidden"}:
            raise RuntimeError("Selftest falhou: aba PLOT_DATA deveria estar oculta.")
        if ws_plot["A1"].value != "TIMESTAMP_TEXT":
            raise RuntimeError("Selftest falhou: cabecalho da aba PLOT_DATA invalido.")
        ws_audit = wb_chk["PLOT_AUDIT"] if "PLOT_AUDIT" in wb_chk.sheetnames else None
        if ws_audit is None:
            raise RuntimeError("Selftest falhou: aba PLOT_AUDIT ausente.")
        ws_time = wb_chk[TIME_AUDIT_SHEET] if TIME_AUDIT_SHEET in wb_chk.sheetnames else None
        if ws_time is None:
            raise RuntimeError("Selftest falhou: aba TIME_AUDIT ausente.")
        if not ws_dash._charts:
            raise RuntimeError("Selftest falhou: grafico ausente no DASHBOARD.")
        expected_series = min(3, len(vars_sel))
        total_plotted = sum(len(chart.series) for chart in ws_dash._charts)
        if total_plotted != expected_series:
            raise RuntimeError(
                f"Selftest falhou: dashboard com {total_plotted} series, esperado {expected_series}."
            )
        audit_yes = sum(1 for row in ws_audit.iter_rows(min_row=2, values_only=True) if str(row[3]) == "YES")
        if audit_yes != expected_series:
            raise RuntimeError(
                f"Selftest falhou: PLOT_AUDIT marcou {audit_yes} series como plotadas, esperado {expected_series}."
            )
        last_chart_end_row = 0
        for chart in ws_dash._charts:
            if chart.legend is None or chart.legend.position not in {"r", "b"}:
                raise RuntimeError("Selftest falhou: legenda do grafico nao foi configurada automaticamente.")
            if bool(chart.legend.overlay):
                raise RuntimeError("Selftest falhou: legenda do grafico nao pode sobrepor o plot.")
            anchor_from = int(getattr(chart.anchor._from, "row", 0)) + 1
            anchor_to = int(getattr(chart.anchor.to, "row", 0)) + 1
            if anchor_from <= last_chart_end_row:
                raise RuntimeError("Selftest falhou: graficos do dashboard ficaram sobrepostos.")
            last_chart_end_row = anchor_to
            num_fmt = getattr(getattr(chart.y_axis, "numFmt", None), "formatCode", "")
            if str(num_fmt) != DASHBOARD_AXIS_NUMFMT:
                raise RuntimeError("Selftest falhou: eixo Y do grafico nao preserva o formato com sinal.")
            for s in chart.series:
                if s.marker is None or s.marker.symbol != "circle":
                    raise RuntimeError("Selftest falhou: marcador de serie nao esta configurado como circle.")
                val_ref = str(getattr(getattr(getattr(s, "val", None), "numRef", None), "f", "") or "")
                if PLOT_DATA_SHEET not in val_ref:
                    raise RuntimeError("Selftest falhou: serie do grafico nao referencia a aba PLOT_DATA.")
    finally:
        wb_chk.close()

    log("XLSX", f"selftest ok: {out}")
    return out


def main():
    ap = argparse.ArgumentParser(description="Andy's Report Runner - Lake -> Recorte -> Dashboard XLSX")
    ap.add_argument("--work-root", default=DEFAULT_WORK_ROOT)
    ap.add_argument("--db", default=None)
    ap.add_argument("--outdir", default=None)
    ap.add_argument("--outname", default=None)

    ap.add_argument("--selftest", action="store_true", help="Executa teste rapido de ponta-a-ponta.")
    ap.add_argument("--equip", action="append", default=None, help="Repita: --equip TR-001 --equip AL-002")
    ap.add_argument("--from", dest="t0", default=None, help='Ex: "2025-01-01 00:00:00"')
    ap.add_argument("--to", dest="t1", default=None, help='Ex: "2025-01-02 23:59:59"')
    ap.add_argument("--var", action="append", default=None, help="Opcional: --var IA --var IB ...")

    ap.add_argument("--agg", default="max", choices=["max", "last"])
    ap.add_argument("--timefloor", default=None, help='Opcional: "15min", "1H", "1D"')
    ap.add_argument("--equip-slots", type=int, default=8)
    ap.add_argument("--var-slots", type=int, default=6)
    ap.add_argument("--max-timestamps", type=int, default=300_000)

    args = ap.parse_args()
    db_path = args.db or get_db_path(args.work_root)
    out_dir = args.outdir or get_export_dir(args.work_root)

    try:
        if args.selftest:
            run_selftest(db_path=db_path, out_dir=out_dir, work_root=args.work_root)
            return

        if not args.equip:
            raise ValueError("Informe ao menos um --equip (ou use --selftest).")
        if not args.t0 or not args.t1:
            raise ValueError("Informe --from e --to (ou use --selftest).")

        run_report(
            work_root=args.work_root,
            db_path=db_path,
            equips=args.equip,
            t0=args.t0,
            t1=args.t1,
            vars_=args.var,
            agg=args.agg,
            time_floor=args.timefloor,
            equip_slots=args.equip_slots,
            var_slots=args.var_slots,
            max_timestamps=args.max_timestamps,
            out_dir=out_dir,
            out_name=args.outname,
        )
    except Exception as e:
        log("ERR", str(e))
        raise


if __name__ == "__main__":
    main()
