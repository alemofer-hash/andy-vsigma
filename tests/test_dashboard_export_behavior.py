"""Workbook contract tests with dynamic runner import."""
from __future__ import annotations

import importlib
import openpyxl
import pandas as pd
import pytest


def _load_runner():
    candidates = [
        "andys_report_runner",
        "xlsx_report_runner",
        "report_runner",
        "desktop_app.andys_report_runner",
        "desktop_app.xlsx_report_runner",
    ]
    for name in candidates:
        try:
            return importlib.import_module(name)
        except Exception:
            continue
    pytest.skip("Nenhum runner XLSX compatível disponível", allow_module_level=True)


runner = _load_runner()
PatamarResult = getattr(runner, "PatamarResult")
PatamarShiftStats = getattr(runner, "PatamarShiftStats")
PatamarWriteSpec = getattr(runner, "PatamarWriteSpec")
construir_bi_excel_v2 = getattr(runner, "construir_bi_excel_v2")
PLOT_DATA_SHEET = getattr(runner, "PLOT_DATA_SHEET")
SHIFT_ORDER = getattr(runner, "SHIFT_ORDER")


def _minimal_long_df(key="||TR-001||IA", var="IA") -> pd.DataFrame:
    return pd.DataFrame({
        "_TS": pd.to_datetime(["2025-01-01 08:00:00", "2025-01-01 09:00:00", "2025-01-01 14:00:00", "2025-01-01 20:00:00"]),
        "_KEY": [key] * 4,
        "_VAL": [1.0, 2.0, 3.0, 4.0],
        "_EQUIP": ["TR-001"] * 4,
        "_VAR": [var] * 4,
        "_CLASSE": ["COR"] * 4,
    })


def _meta() -> dict:
    return {
        "equips": "TR-001",
        "t0": "2025-01-01 00:00:00",
        "t1": "2025-01-01 23:59:59",
        "agg": "max",
        "time_floor": "(none)",
        "generated_at": "2025-01-01 00:00:00",
        "period_label": "2025-01-01",
        "ts_min": "2025-01-01 08:00:00",
        "ts_max": "2025-01-01 20:00:00",
    }


def _patamar_result(p_text="1.0/2.0", q_text="0.5/1.0"):
    return PatamarResult(
        shifts=[PatamarShiftStats(shift=s, p_min=1.0, p_max=2.0, q_min=0.5, q_max=1.0, p_text=p_text, q_text=q_text) for s in SHIFT_ORDER],
        p_key_used="||TR-001||P",
        q_key_used="||TR-001||Q",
        warnings=[],
    )


def test_plot_data_hidden_default(tmp_path):
    out = str(tmp_path / "out.xlsx")
    construir_bi_excel_v2(_minimal_long_df(), out, _meta())
    wb = openpyxl.load_workbook(out)
    assert PLOT_DATA_SHEET in wb.sheetnames
    assert wb[PLOT_DATA_SHEET].sheet_state in {"hidden", "veryHidden"}
    wb.close()


def test_patamar_summary_filled_when_template_has_sheet(tmp_path):
    tpl = str(tmp_path / "tpl.xlsx")
    wb = openpyxl.Workbook()
    wb.create_sheet("PATAMAR - Resumo")
    wb.save(tpl)
    out = str(tmp_path / "out.xlsx")
    construir_bi_excel_v2(_minimal_long_df(), out, _meta(), template_path=tpl, patamar_result=_patamar_result(), patamar_write_spec=PatamarWriteSpec())
    wb2 = openpyxl.load_workbook(out, data_only=True)
    ws = wb2["PATAMAR - Resumo"]
    for row in range(6, 10):
        assert isinstance(ws[f"L{row}"].value, str)
        assert isinstance(ws[f"M{row}"].value, str)
    wb2.close()
