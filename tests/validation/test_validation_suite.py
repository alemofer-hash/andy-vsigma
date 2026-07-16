from __future__ import annotations

from pathlib import Path
from typing import Dict, List
import uuid

import duckdb
import openpyxl
import pandas as pd
import pytest

import andys_table_app as app
from andys_indexer import get_variable_cols, indexar_tudo
from andys_report_runner import aggregate_long, construir_bi_excel_multi_equip_multi_var_long
from db.query_builder import (
    build_distinct_query,
    build_filters,
    build_pagination,
    build_ponto_query,
    build_vars_for_pontos_query,
)
from xlsx_selection import pontos_to_xlsx_selection


# --- NEW: fixture integrada para validar ingestao + lake + duckdb em ambiente temporario ---
@pytest.fixture(scope="module")
def validation_workspace() -> Dict[str, Path]:
    base = (Path.cwd() / ".validation_tmp" / "tests" / f"andys_validation_{uuid.uuid4().hex[:10]}").resolve()
    source = base / "source"
    allowed = base / "allowed"
    work = allowed / "work"
    reports = base / "reports"

    source.mkdir(parents=True, exist_ok=True)
    work.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    csv_file = source / "Parametros eletricos - 01_2025 - Todas SEs.csv"
    csv_file.write_text(
        "E3TIMESTAMP;SE;BAY;EQUIPAMENTO;TERMINAL;IA;IB;VAB;CUSTOM_VAR\n"
        "01/01/2025 00:00:00;SE1;BAY_A;TR-1;Terminal1;10,0;11,0;13,8;2,5\n"
        "01/01/2025 00:00:00;SE1;BAY_B;TR-1;Terminal2;12,0;13,0;14,1;3,5\n",
        encoding="utf-8",
    )

    xlsx_file = source / "Parametros eletricos - 02_2025 - Todas SEs.xlsx"
    xlsx_df = pd.DataFrame(
        {
            "E3TIMESTAMP": ["01/02/2025 00:00:00", "01/02/2025 00:15:00"],
            "SE": ["SE1", "SE1"],
            "BAY": ["BAY_A", "BAY_B"],
            "EQUIPAMENTO": ["TR-1", "TR-1"],
            "TERMINAL": ["Terminal1", "Terminal2"],
            "IA": ["15,5", "16,5"],
            "IB": ["17,5", "18,5"],
            "VAB": ["13,2", "13,6"],
        }
    )
    xlsx_df.to_excel(xlsx_file, index=False)

    indexar_tudo(source_root=str(source), work_root=str(work), allowed_root=str(allowed))

    lake = work / "ANDYS_LAKE"
    db_path = lake / "andys.duckdb"
    assert db_path.exists(), "DuckDB deveria ter sido gerado durante a indexacao."

    return {
        "base": base,
        "source": source,
        "allowed": allowed,
        "work": work,
        "lake": lake,
        "db_path": db_path,
        "reports": reports,
    }


@pytest.mark.validation
@pytest.mark.smoke
def test_smoke_imports_and_entrypoints() -> None:
    assert callable(indexar_tudo)
    assert callable(app.query_page)
    assert callable(build_filters)
    assert callable(construir_bi_excel_multi_equip_multi_var_long)


@pytest.mark.validation
@pytest.mark.integration
def test_ingestion_and_duckdb_availability(validation_workspace: Dict[str, Path]) -> None:
    db_path = validation_workspace["db_path"]
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        tables = {str(r[0]).lower() for r in con.execute("SHOW TABLES;").fetchall()}
        assert "medicoes" in tables
        row_count = int(con.execute("SELECT COUNT(*) FROM medicoes;").fetchone()[0])
        assert row_count > 0
        sample = con.execute("SELECT timestamp, SE, BAY, EQUIPAMENTO, TERMINAL, ponto_id, var, valor FROM medicoes LIMIT 1;").fetchone()
        assert sample is not None
    finally:
        con.close()


@pytest.mark.validation
@pytest.mark.integration
def test_dynamic_variable_discovery_from_canonical_columns() -> None:
    df = pd.DataFrame(
        {
            "TIMESTAMP": ["2025-01-01 00:00:00"],
            "SE": ["SE1"],
            "BAY": ["BAY_A"],
            "EQUIPAMENTO": ["TR-1"],
            "TERMINAL": ["Terminal1"],
            "IA": ["10,0"],
            "IB": ["11,0"],
            "VAB": ["13,8"],
            "CUSTOM_VAR": ["99,9"],
        }
    )
    cols = get_variable_cols(df)
    assert {"IA", "IB", "VAB", "CUSTOM_VAR"}.issubset(set(cols))


@pytest.mark.validation
@pytest.mark.integration
def test_hierarchical_resolution_and_point_scoped_vars(validation_workspace: Dict[str, Path]) -> None:
    db_path = validation_workspace["db_path"]
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        ano, mes = con.execute(
            """
            SELECT ano, mes
            FROM medicoes
            WHERE COALESCE(CAST(SE AS VARCHAR), '') <> ''
            ORDER BY ano DESC, mes DESC
            LIMIT 1
            """
        ).fetchone()
        assert ano is not None and mes is not None

        sql_se, p_se = build_distinct_query(target_col="SE", ano=int(ano), mes=int(mes))
        se_options = [r[0] for r in con.execute(sql_se, p_se).fetchall()]
        assert se_options

        sql_bay, p_bay = build_distinct_query(target_col="BAY", ano=int(ano), mes=int(mes), se_sel=[se_options[0]], include_empty=True)
        bay_options = [r[0] for r in con.execute(sql_bay, p_bay).fetchall()]
        assert bay_options

        sql_eq, p_eq = build_distinct_query(
            target_col="EQUIPAMENTO",
            ano=int(ano),
            mes=int(mes),
            se_sel=[se_options[0]],
            bay_sel=[bay_options[0]],
            include_empty=True,
        )
        equip_options = [r[0] for r in con.execute(sql_eq, p_eq).fetchall()]
        assert equip_options

        sql_term, p_term = build_distinct_query(
            target_col="TERMINAL",
            ano=int(ano),
            mes=int(mes),
            se_sel=[se_options[0]],
            bay_sel=[bay_options[0]],
            equipamento_sel=[equip_options[0]],
            include_empty=True,
        )
        terminal_options = [r[0] for r in con.execute(sql_term, p_term).fetchall()]
        assert terminal_options

        sql_pts, p_pts = build_ponto_query(
            ano=int(ano),
            mes=int(mes),
            se_sel=[se_options[0]],
            bay_sel=[bay_options[0]],
            equipamento_sel=[equip_options[0]],
            terminal_sel=[terminal_options[0]],
            limit=200,
        )
        pontos = [str(r[0]) for r in con.execute(sql_pts, p_pts).fetchall() if r[0] is not None]
        assert pontos

        sql_vars, p_vars = build_vars_for_pontos_query(ano=int(ano), mes=int(mes), ponto_ids_sel=pontos)
        vars_by_points = [str(r[0]) for r in con.execute(sql_vars, p_vars).fetchall()]
        assert vars_by_points
        assert "IA" in vars_by_points
    finally:
        con.close()


@pytest.mark.validation
@pytest.mark.integration
def test_main_query_page_path_and_pagination(validation_workspace: Dict[str, Path]) -> None:
    db_path = str(validation_workspace["db_path"])
    con = duckdb.connect(db_path, read_only=True)
    try:
        ano, mes = con.execute("SELECT MAX(ano), MAX(mes) FROM medicoes;").fetchone()
        ponto_ids = [str(r[0]) for r in con.execute("SELECT DISTINCT ponto_id FROM medicoes LIMIT 2;").fetchall()]
    finally:
        con.close()

    where_sql, where_params = build_filters(
        equips_selected=[],
        equip_like=None,
        vars_sel=["IA"],
        ponto_ids_sel=ponto_ids,
        ano=int(ano),
        mes=int(mes),
        t0=None,
        t1=None,
    )
    pagination_sql, pagination_params = build_pagination(limit=50, offset=0)
    total, page_df = app.query_page(
        db_path=db_path,
        where_sql=where_sql,
        where_params=where_params,
        pagination_sql=pagination_sql,
        pagination_params=pagination_params,
        order_sql="ORDER BY timestamp ASC",
    )
    assert total >= 0
    assert isinstance(page_df, pd.DataFrame)
    assert len(page_df) <= 50


@pytest.mark.validation
@pytest.mark.integration
def test_export_csv_and_xlsx_dashboard(validation_workspace: Dict[str, Path]) -> None:
    db_path = str(validation_workspace["db_path"])
    reports_dir = validation_workspace["reports"]

    con = duckdb.connect(db_path, read_only=True)
    try:
        long_df = con.execute(
            """
            SELECT timestamp, SE, BAY, TERMINAL, equip_id, var, classe, valor
            FROM medicoes
            WHERE var = 'IA'
            ORDER BY timestamp ASC
            """
        ).df()
    finally:
        con.close()

    assert not long_df.empty
    if "TIMESTAMP" in long_df.columns and "timestamp" not in long_df.columns:
        long_df = long_df.rename(columns={"TIMESTAMP": "timestamp"})
    csv_out = reports_dir / "validation_export.csv"
    long_df.to_csv(csv_out, index=False)
    assert csv_out.exists()
    loaded_csv = pd.read_csv(csv_out)
    assert not loaded_csv.empty

    df_agg = aggregate_long(long_df, agg="max", time_floor=None)
    assert not df_agg.empty

    xlsx_out = reports_dir / "validation_dashboard.xlsx"
    out_file, warnings = construir_bi_excel_multi_equip_multi_var_long(
        long_df=df_agg,
        xlsx_out=str(xlsx_out),
        report_meta={"scope": "validation"},
        selected_pairs=[("TR-1", "IA")],
    )
    assert Path(out_file).exists()
    assert isinstance(warnings, list)

    wb = openpyxl.load_workbook(out_file, data_only=False)
    try:
        assert "VIEW_MX" in wb.sheetnames
        assert "DASHBOARD" in wb.sheetnames
        ws_view = wb["VIEW_MX"]
        ws_dash = wb["DASHBOARD"]
        assert ws_view["A4"].value == "TIMESTAMP_TEXT"
        assert isinstance(ws_view["A8"].value, str)
        assert ws_dash._charts, "Dashboard precisa conter ao menos um grafico."
        assert len(ws_dash._charts[0].series) >= 2
    finally:
        wb.close()


@pytest.mark.validation
@pytest.mark.integration
def test_xlsx_selection_compatibility_from_points(validation_workspace: Dict[str, Path]) -> None:
    db_path = str(validation_workspace["db_path"])
    con = duckdb.connect(db_path, read_only=True)
    try:
        pontos = [str(r[0]) for r in con.execute("SELECT DISTINCT ponto_id FROM medicoes WHERE EQUIPAMENTO = 'TR-1' ORDER BY ponto_id;").fetchall()]
    finally:
        con.close()

    selections = pontos_to_xlsx_selection(pontos, ["IA", "IB", "IA"])
    assert "TR-1" in selections
    assert selections["TR-1"] == ["IA", "IB"]
