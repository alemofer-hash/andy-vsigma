from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from .cadence import infer_cadence
from .key_parser import is_feeder_name, parse_andy_key
from .models import CURRENT_VARIABLES, FeederProfile

REQUIRED_COLUMNS = ("TIMESTAMP", "KEY", "VAL", "EQUIP", "VAR", "CLASSE", "TSKEY")


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_current_variables(values: Iterable[object] | None = None) -> tuple[str, ...]:
    if values is None:
        return tuple(sorted(CURRENT_VARIABLES))
    normalized = tuple(
        dict.fromkeys(
            str(value or "").strip().upper()
            for value in values
            if str(value or "").strip().upper() in CURRENT_VARIABLES
        )
    )
    return normalized or tuple(sorted(CURRENT_VARIABLES))


def read_workbook_currents(
    path: str | Path,
    *,
    current_variables: Iterable[object] | None = None,
) -> tuple[pd.DataFrame, dict[str, Any], list[FeederProfile]]:
    source_path = Path(path)
    source_hash = hash_file(source_path)
    alias = source_path.name
    selected_vars = set(normalize_current_variables(current_variables))
    wb = load_workbook(source_path, read_only=True, data_only=True, keep_links=False)
    try:
        if "DADOS_AGG" not in wb.sheetnames:
            raise ValueError("DADOS_AGG_sheet_required")
        ws = wb["DADOS_AGG"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        missing = [name for name in REQUIRED_COLUMNS if name not in header]
        if missing:
            raise ValueError("missing_columns:" + ",".join(missing))
        idx = {name: header.index(name) for name in REQUIRED_COLUMNS}
        records: list[dict[str, Any]] = []
        total_rows = 0
        var_counts: dict[str, int] = {}
        parse_errors = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            raw_var = str(row[idx["VAR"]] or "").strip().upper()
            var_counts[raw_var] = var_counts.get(raw_var, 0) + 1
            if raw_var not in selected_vars:
                continue
            try:
                parsed = parse_andy_key(str(row[idx["KEY"]] or ""))
            except ValueError:
                parse_errors += 1
                continue
            if parsed.variable.upper() not in selected_vars:
                continue
            try:
                value = float(row[idx["VAL"]])
            except (TypeError, ValueError):
                value = float("nan")
            records.append(
                {
                    "source_alias": alias,
                    "source_hash": source_hash,
                    "timestamp": pd.to_datetime(row[idx["TIMESTAMP"]], errors="coerce"),
                    "key": parsed.canonical,
                    "se": parsed.se,
                    "feeder": parsed.feeder,
                    "equipment": parsed.equipment,
                    "terminal": parsed.terminal,
                    "variable": parsed.variable,
                    "value": value,
                    "raw_equip": row[idx["EQUIP"]],
                    "classe": row[idx["CLASSE"]],
                    "tskey": row[idx["TSKEY"]],
                    "is_feeder": is_feeder_name(parsed.feeder),
                }
            )
    finally:
        wb.close()
    df = pd.DataFrame.from_records(records)
    if not df.empty:
        df = df.dropna(subset=["timestamp"]).sort_values(["se", "feeder", "timestamp"]).reset_index(drop=True)
    manifest = {
        "source_alias": alias,
        "source_hash": source_hash,
        "sheet": "DADOS_AGG",
        "total_rows": total_rows,
        "current_rows": int(len(df)),
        "selected_current_variables": sorted(selected_vars),
        "var_counts": var_counts,
        "parse_errors": parse_errors,
        "columns": list(REQUIRED_COLUMNS),
        "read_only": True,
        "content_opened": "metadata_and_DADOS_AGG_values_for_selected_current_variables_only",
    }
    profiles = profile_feeders(df)
    return df, manifest, profiles


def read_workbook_ib(path: str | Path) -> tuple[pd.DataFrame, dict[str, Any], list[FeederProfile]]:
    return read_workbook_currents(path, current_variables=("IB",))


def profile_feeders(df: pd.DataFrame) -> list[FeederProfile]:
    profiles: list[FeederProfile] = []
    if df.empty:
        return profiles
    for key, part in df.groupby("key"):
        ordered = part.sort_values("timestamp")
        cadence = infer_cadence(ordered["timestamp"])
        values = ordered["value"].dropna()
        constant = bool(values.nunique(dropna=True) <= 1) if not values.empty else True
        first = ordered.iloc[0]
        profiles.append(
            FeederProfile(
                source_alias=str(first["source_alias"]),
                source_hash=str(first["source_hash"]),
                key=str(key),
                se=str(first["se"]),
                feeder=str(first["feeder"]),
                equipment=str(first["equipment"]),
                terminal=str(first["terminal"]),
                variable=str(first["variable"]),
                rows=int(len(ordered)),
                period_start=cadence.coverage_start,
                period_end=cadence.coverage_end,
                cadence_seconds=cadence.cadence_seconds,
                gap_count=cadence.gap_count,
                duplicate_count=cadence.duplicate_count,
                missing_count=int(ordered["value"].isna().sum()),
                constant_series=constant,
            )
        )
    return profiles
